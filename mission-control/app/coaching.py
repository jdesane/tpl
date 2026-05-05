"""
Phase 15 — Coaching Platform module.

Surfaces:
  - Coaching client list / create / update / delete
  - Business Plan editor (Budget Model + Economic Model + Lead Gen Model)
  - Computed numbers endpoint (auditable: every value carries its formula + inputs)

The `db` callable is injected by main.py so we inherit the workspace-scoping wrapper
without circular imports. `supabase` is the raw client for child tables that scope
via FK chain (budget_models, economic_models, etc.) and don't carry workspace_id.
"""
from fastapi import APIRouter, HTTPException, Request
from pydantic import BaseModel
from typing import Optional, Callable, Any
from datetime import date, datetime

router = APIRouter(prefix="/api/coaching", tags=["coaching"])

# Injected by main.py via setup() ─ avoids circular imports.
_db: Optional[Callable[[str], Any]] = None
_supabase: Any = None


def setup(db_callable, supabase_client):
    """Called from main.py after `db` and `supabase` are defined."""
    global _db, _supabase
    _db = db_callable
    _supabase = supabase_client


# ════════════════════════════════════════════════════════════
# Pydantic models
# ════════════════════════════════════════════════════════════

class CoachingClientIn(BaseModel):
    # Existing-contact path: just pass lead_id
    lead_id: Optional[int] = None
    # New-contact path: pass these to create the lead AND coaching_client in one call
    new_contact: Optional[dict] = None  # {first_name, last_name, email, phone, current_brokerage}
    # Coaching metadata (all optional at creation; can be filled in later)
    brokerage: Optional[str] = None
    lpt_comp_plan: Optional[str] = None
    license_date: Optional[str] = None
    market_city: Optional[str] = None
    market_state: Optional[str] = None
    avg_sale_price: Optional[float] = None
    avg_commission_rate: Optional[float] = None
    call_cadence: Optional[str] = "WEEKLY"
    coaching_start_date: Optional[str] = None
    notes: Optional[str] = None
    # If True, auto-provision portal login + send invite email immediately after creating the client.
    # The MC modal defaults this to True so "Invite Client" actually sends an invite.
    send_portal_invite: Optional[bool] = False


class CoachingClientUpdate(BaseModel):
    brokerage: Optional[str] = None
    lpt_comp_plan: Optional[str] = None
    license_date: Optional[str] = None
    market_city: Optional[str] = None
    market_state: Optional[str] = None
    avg_sale_price: Optional[float] = None
    avg_commission_rate: Optional[float] = None
    call_cadence: Optional[str] = None
    coaching_start_date: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None


class BusinessPlanUpdate(BaseModel):
    gci_target: Optional[float] = None
    notes: Optional[str] = None


class BudgetModelUpdate(BaseModel):
    paid_to_brokerage: Optional[float] = None
    referrals_to_you_count: Optional[int] = None
    referrals_avg_commission: Optional[float] = None
    referrals_split_pct: Optional[float] = None
    seller_specialist_split_pct: Optional[float] = None
    buyer_specialist_split_pct: Optional[float] = None
    operating_expenses: Optional[dict] = None
    charity_pct: Optional[float] = None
    retirement_pct: Optional[float] = None
    income_tax_pct: Optional[float] = None
    personal_expenses: Optional[dict] = None
    avg_net_commission_per_close: Optional[float] = None
    surplus_allocation: Optional[list] = None


class EconomicModelUpdate(BaseModel):
    seller_pct: Optional[float] = None
    seller_avg_sale_price: Optional[float] = None
    buyer_avg_sale_price: Optional[float] = None
    commission_rate: Optional[float] = None
    listings_close_pct: Optional[float] = None
    buyers_close_pct: Optional[float] = None
    listing_appt_to_list_pct: Optional[float] = None
    buyer_appt_to_work_pct: Optional[float] = None
    avg_days_on_market: Optional[int] = None
    avg_buyer_working_days: Optional[int] = None


# ════════════════════════════════════════════════════════════
# Default operating-expense / personal-expense scaffolds
# ════════════════════════════════════════════════════════════
# These are line-item TEMPLATES — agents can add/remove rows from the UI.
# JSONB structure: { "Category > Line Item": amount }

DEFAULT_OPERATING_EXPENSES = {
    "Salaries > Admin": 0,
    "Salaries > Marketing": 0,
    "Lead Gen Marketing > Advertising in media": 0,
    "Lead Gen Marketing > Photography / staging": 0,
    "Lead Gen Marketing > Websites / internet leads / social clicks": 0,
    "Lead Gen Marketing > Direct mail postcards": 0,
    "Lead Gen Marketing > Signs / brochure boxes": 0,
    "Lead Gen Marketing > Name tags / shirts / car magnets": 0,
    "Lead Gen Marketing > Sponsoring festivals / teams": 0,
    "Lead Gen Prospecting > Open houses": 0,
    "Lead Gen Prospecting > Client parties": 0,
    "Lead Gen Prospecting > Networking events": 0,
    "Lead Gen Prospecting > Teaching / speaking": 0,
    "Lead Gen Prospecting > Builders / FSBOs direct": 0,
    "Lead Gen Prospecting > Door to door": 0,
    "Occupancy > Rent": 0,
    "Technology > Apps / software / subscriptions": 0,
    "Technology > Phone": 0,
    "Operations > Supplies": 0,
    "Operations > Education / Dues / Coaching / Travel": 0,
    "Operations > Equipment (long-term use)": 0,
    "Auto > Gas": 0,
    "Auto > Maintenance": 0,
    "Auto > Insurance": 0,
}

DEFAULT_PERSONAL_EXPENSES = {
    "House": 0,
    "Food": 0,
    "Recreation / Entertainment": 0,
    "Credit Cards": 0,
    "Car": 0,
    "Other Regular Payments": 0,
    "Miscellaneous": 0,
}

# Default LPT cap based on comp plan
LPT_CAP_BUSINESS_BUILDER = 5000
LPT_CAP_BROKERAGE_PARTNER = 15000
LPT_TC_FEE_PER_TXN = 195       # Transaction-coordinator fee on every closed deal
LPT_BB_PER_TXN_FEE = 500       # Business Builder per-transaction office fee until cap


def lpt_plan_details(comp_plan: str) -> dict:
    """Returns the LPT comp plan constants. Used to auto-populate the wizard
    so agents don't have to type splits / caps / fees that are already known."""
    if comp_plan == "BROKERAGE_PARTNER":
        return {
            "comp_plan": "BROKERAGE_PARTNER",
            "label": "Brokerage Partner",
            "split_to_office_pct": 0.20,        # 80/20 split
            "cap": LPT_CAP_BROKERAGE_PARTNER,
            "per_txn_fee": 0,
            "tc_fee_per_txn": LPT_TC_FEE_PER_TXN,
            "royalty_pct": 0,
        }
    elif comp_plan == "BUSINESS_BUILDER":
        return {
            "comp_plan": "BUSINESS_BUILDER",
            "label": "Business Builder",
            "split_to_office_pct": 0,           # Flat-fee model, not split
            "cap": LPT_CAP_BUSINESS_BUILDER,
            "per_txn_fee": LPT_BB_PER_TXN_FEE,
            "tc_fee_per_txn": LPT_TC_FEE_PER_TXN,
            "royalty_pct": 0,
        }
    return None


def compute_gci_from_net_income(
    net_income_goal: float,
    avg_gci_per_deal: float,
    comp_plan: str,
    opex_pct: float = 0.25,
) -> dict:
    """Iteratively solve for the GCI target needed to hit a net-income-before-taxes goal,
    given the agent's avg deal size and LPT comp plan. Returns the implied closings,
    cost-of-sale breakdown, and operating expense estimate.

    Uses fixed-point iteration because COGS depends on closings (cap behavior + TC fee
    × txns), and closings depend on GCI = net + COGS + OpEx.
    """
    if net_income_goal <= 0 or avg_gci_per_deal <= 0:
        return None
    plan = lpt_plan_details(comp_plan) or lpt_plan_details("BROKERAGE_PARTNER")
    cap = plan["cap"]
    split_pct = plan["split_to_office_pct"]
    per_txn_fee = plan["per_txn_fee"]
    tc_fee = plan["tc_fee_per_txn"]

    # Initial guess
    closings = max(1, net_income_goal / avg_gci_per_deal)
    for _ in range(40):
        # Office split / per-txn fee, capped
        if split_pct > 0:
            office_uncapped = avg_gci_per_deal * split_pct * closings
        else:
            office_uncapped = per_txn_fee * closings
        office_total = min(cap, office_uncapped)
        tc_total = tc_fee * closings
        cogs = office_total + tc_total

        gci_estimate = closings * avg_gci_per_deal
        opex_total = opex_pct * gci_estimate

        required_gci = net_income_goal + cogs + opex_total
        new_closings = required_gci / avg_gci_per_deal
        if abs(new_closings - closings) < 0.01:
            closings = new_closings
            break
        closings = new_closings

    return {
        "gci_target": round(closings * avg_gci_per_deal, 2),
        "closings": round(closings, 1),
        "office_cogs": round(office_total, 2),
        "tc_cogs": round(tc_total, 2),
        "total_cogs": round(office_total + tc_total, 2),
        "opex_estimate": round(opex_total, 2),
        "net_income_goal": net_income_goal,
        "avg_gci_per_deal": avg_gci_per_deal,
        "avg_net_per_close": round((closings * avg_gci_per_deal - office_total - tc_total - opex_total) / closings, 2) if closings > 0 else 0,
    }


# ════════════════════════════════════════════════════════════
# Helpers
# ════════════════════════════════════════════════════════════

def _ws(req: Request) -> int:
    return getattr(req.state, "workspace_id", 1) or 1


def _normalize_pct(v):
    """Accept percentages as either 0.30 or 30 — snap to 0–1 range."""
    if v is None:
        return None
    f = float(v)
    return f / 100.0 if f > 1.0 else f


def _enrich_client(row: dict) -> dict:
    """Attach lead-name fields onto a coaching_client row."""
    if not row:
        return row
    lead_id = row.get("lead_id")
    if lead_id:
        try:
            r = _supabase.table("leads").select("first_name,last_name,name,email,phone,current_brokerage").eq("id", lead_id).single().execute()
            if r.data:
                row["lead"] = r.data
        except Exception:
            row["lead"] = None
    return row


def _ensure_business_plan(client_id: int, year: int, workspace_id: int) -> dict:
    """Find or create the business_plan + budget_model + economic_model + lead_gen_model rows."""
    bp_resp = _supabase.table("business_plans").select("*").eq("coaching_client_id", client_id).eq("year", year).execute()
    if bp_resp.data:
        plan = bp_resp.data[0]
    else:
        ins = _supabase.table("business_plans").insert({
            "workspace_id": workspace_id,
            "coaching_client_id": client_id,
            "year": year,
            "gci_target": 0,
        }).execute()
        plan = ins.data[0]

    plan_id = plan["id"]

    # Budget model — auto-create with seeded operating/personal expense scaffolds
    bm_resp = _supabase.table("budget_models").select("*").eq("business_plan_id", plan_id).execute()
    if bm_resp.data:
        budget = bm_resp.data[0]
    else:
        # Default paid_to_brokerage from the client's comp plan, if set
        client = _supabase.table("coaching_clients").select("lpt_comp_plan").eq("id", client_id).single().execute().data or {}
        cap = 0
        if client.get("lpt_comp_plan") == "BROKERAGE_PARTNER":
            cap = LPT_CAP_BROKERAGE_PARTNER
        elif client.get("lpt_comp_plan") == "BUSINESS_BUILDER":
            cap = LPT_CAP_BUSINESS_BUILDER
        ins = _supabase.table("budget_models").insert({
            "business_plan_id": plan_id,
            "paid_to_brokerage": cap,
            "operating_expenses": DEFAULT_OPERATING_EXPENSES,
            "personal_expenses": DEFAULT_PERSONAL_EXPENSES,
        }).execute()
        budget = ins.data[0]

    # Economic model
    em_resp = _supabase.table("economic_models").select("*").eq("business_plan_id", plan_id).execute()
    if em_resp.data:
        economic = em_resp.data[0]
    else:
        # Default sale prices from the client's avg_sale_price (same value for both sides until split)
        client = _supabase.table("coaching_clients").select("avg_sale_price,avg_commission_rate").eq("id", client_id).single().execute().data or {}
        asp = float(client.get("avg_sale_price") or 0)
        rate = float(client.get("avg_commission_rate") or 0.025)
        ins = _supabase.table("economic_models").insert({
            "business_plan_id": plan_id,
            "seller_avg_sale_price": asp,
            "buyer_avg_sale_price": asp,
            "commission_rate": rate,
        }).execute()
        economic = ins.data[0]

    # Lead Gen model
    lg_resp = _supabase.table("lead_gen_models").select("*").eq("business_plan_id", plan_id).execute()
    if lg_resp.data:
        leadgen = lg_resp.data[0]
    else:
        ins = _supabase.table("lead_gen_models").insert({
            "business_plan_id": plan_id,
        }).execute()
        leadgen = ins.data[0]

    return {
        "plan": plan,
        "budget_model": budget,
        "economic_model": economic,
        "lead_gen_model": leadgen,
    }


# ════════════════════════════════════════════════════════════
# Math: Economic Model derivations
# ════════════════════════════════════════════════════════════
# Mirrors the MREA workbook:
#   - Seller revenue = GCI × seller %
#   - Buyer revenue  = GCI − seller revenue
#   - Volume         = revenue / commission rate
#   - Houses sold    = volume / avg sale price
#   - Listings taken = houses sold (sellers) / % listings close
#   - Buyers shown   = houses sold (buyers) / % buyers close
#   - Listing appts  = listings taken / % appts that list
#   - Buyer consults = buyers shown / % buyer appts work
#   - Per week       = annual / 48  (NOT 52 — accounts for vacation/holidays per legacy spreadsheet)
#   - Active listings to carry = houses sold sellers / (365 / DOM)
#   - Ready buyers to work     = houses sold buyers  / (365 / buyer working time)

def _safe_div(a, b):
    if b is None or b == 0:
        return 0
    return a / b


def calc_economic(gci_target: float, em: dict) -> dict:
    seller_pct = float(em.get("seller_pct") or 0.5)
    sasp = float(em.get("seller_avg_sale_price") or 0)
    basp = float(em.get("buyer_avg_sale_price") or 0)
    rate = float(em.get("commission_rate") or 0.025)
    list_close = float(em.get("listings_close_pct") or 0.85)
    buyer_close = float(em.get("buyers_close_pct") or 0.80)
    list_appt_pct = float(em.get("listing_appt_to_list_pct") or 0.65)
    buyer_appt_pct = float(em.get("buyer_appt_to_work_pct") or 0.75)
    dom = float(em.get("avg_days_on_market") or 60)
    buyer_days = float(em.get("avg_buyer_working_days") or 60)

    seller_rev = gci_target * seller_pct
    buyer_rev = gci_target - seller_rev

    seller_volume = _safe_div(seller_rev, rate)
    buyer_volume = _safe_div(buyer_rev, rate)

    sellers_closed = _safe_div(seller_volume, sasp) if sasp > 0 else 0
    buyers_closed = _safe_div(buyer_volume, basp) if basp > 0 else 0

    listings_taken = _safe_div(sellers_closed, list_close)
    buyers_worked = _safe_div(buyers_closed, buyer_close)

    listing_appts_yr = _safe_div(listings_taken, list_appt_pct)
    buyer_consults_yr = _safe_div(buyers_worked, buyer_appt_pct)

    listings_per_wk = listing_appts_yr / 48.0
    buyers_per_wk = buyer_consults_yr / 48.0

    active_listings_carry = _safe_div(sellers_closed, _safe_div(365, dom)) if dom > 0 else 0
    ready_buyers = _safe_div(buyers_closed, _safe_div(365, buyer_days)) if buyer_days > 0 else 0

    return {
        "seller_revenue":          {"value": round(seller_rev, 2),       "formula": "GCI Target × Seller %"},
        "buyer_revenue":           {"value": round(buyer_rev, 2),        "formula": "GCI Target − Seller Revenue"},
        "seller_volume":           {"value": round(seller_volume, 2),    "formula": "Seller Revenue ÷ Commission Rate"},
        "buyer_volume":            {"value": round(buyer_volume, 2),     "formula": "Buyer Revenue ÷ Commission Rate"},
        "sellers_closed":          {"value": round(sellers_closed, 2),   "formula": "Seller Volume ÷ Seller Avg Sale Price"},
        "buyers_closed":           {"value": round(buyers_closed, 2),    "formula": "Buyer Volume ÷ Buyer Avg Sale Price"},
        "houses_sold_total":       {"value": round(sellers_closed + buyers_closed, 2), "formula": "Sellers Closed + Buyers Closed"},
        "listings_taken":          {"value": round(listings_taken, 2),   "formula": "Sellers Closed ÷ % Listings Close"},
        "buyers_shown":            {"value": round(buyers_worked, 2),    "formula": "Buyers Closed ÷ % Buyers Close"},
        "listing_appts_annual":    {"value": round(listing_appts_yr, 1), "formula": "Listings Taken ÷ % Appts that List"},
        "buyer_consults_annual":   {"value": round(buyer_consults_yr, 1),"formula": "Buyers Shown ÷ % Buyer Appts that Work"},
        "listing_appts_per_week":  {"value": round(listings_per_wk, 2),  "formula": "Listing Appts Annual ÷ 48"},
        "buyer_consults_per_week": {"value": round(buyers_per_wk, 2),    "formula": "Buyer Consults Annual ÷ 48"},
        "active_listings_to_carry":{"value": round(active_listings_carry, 1), "formula": "Sellers Closed ÷ (365 ÷ Avg DOM)"},
        "ready_buyers":            {"value": round(ready_buyers, 1),     "formula": "Buyers Closed ÷ (365 ÷ Avg Buyer Working Days)"},
    }


# ════════════════════════════════════════════════════════════
# Math: Budget Model derivations
# ════════════════════════════════════════════════════════════

def calc_budget(gci_target: float, bm: dict, econ_calc: dict) -> dict:
    paid_to_brokerage = float(bm.get("paid_to_brokerage") or 0)
    ref_count = float(bm.get("referrals_to_you_count") or 0)
    ref_avg = float(bm.get("referrals_avg_commission") or 0)
    ref_split = float(bm.get("referrals_split_pct") or 0.25)
    seller_split = float(bm.get("seller_specialist_split_pct") or 0)  # 0 means no team
    buyer_split = float(bm.get("buyer_specialist_split_pct") or 0)

    seller_rev = econ_calc["seller_revenue"]["value"]
    buyer_rev = econ_calc["buyer_revenue"]["value"]

    referral_fees = ref_count * ref_avg * ref_split
    seller_specialist_cost = seller_rev * seller_split
    buyer_specialist_cost = buyer_rev * buyer_split
    total_cost_of_sale = paid_to_brokerage + referral_fees + seller_specialist_cost + buyer_specialist_cost

    gross_profit = gci_target - total_cost_of_sale

    op_exp = bm.get("operating_expenses") or {}
    if isinstance(op_exp, str):
        try:
            import json as _j
            op_exp = _j.loads(op_exp)
        except Exception:
            op_exp = {}
    total_op_exp = sum(float(v or 0) for v in op_exp.values()) if isinstance(op_exp, dict) else 0

    net_income = gross_profit - total_op_exp

    charity_pct = float(bm.get("charity_pct") or 0)
    retirement_pct = float(bm.get("retirement_pct") or 0)
    income_tax_pct = float(bm.get("income_tax_pct") or 0)

    charity_amt = net_income * charity_pct
    retirement_amt = net_income * retirement_pct
    taxable_income = net_income - charity_amt - retirement_amt
    tax_amt = taxable_income * income_tax_pct
    take_home = taxable_income - tax_amt
    monthly_budget = take_home / 12.0

    # Personal expenses
    pers = bm.get("personal_expenses") or {}
    if isinstance(pers, str):
        try:
            import json as _j
            pers = _j.loads(pers)
        except Exception:
            pers = {}
    monthly_personal = sum(float(v or 0) for v in pers.values()) if isinstance(pers, dict) else 0
    annual_personal = monthly_personal * 12.0

    # Survival number — total annual need = personal annual + business op exp + (each × 30% for taxes)
    annual_need_pre_tax = annual_personal + total_op_exp
    annual_need_w_tax = annual_need_pre_tax * 1.30
    avg_net_close = float(bm.get("avg_net_commission_per_close") or 0)
    survival_closings = _safe_div(annual_need_w_tax, avg_net_close)
    closings_total = econ_calc["houses_sold_total"]["value"]
    closings_beyond_survival = closings_total - survival_closings
    surplus = closings_beyond_survival * avg_net_close * (1 - income_tax_pct)

    return {
        "referral_fees":           {"value": round(referral_fees, 2),         "formula": "# Referrals to You × Avg Referral Commission × Referral Split %"},
        "seller_specialist_cost":  {"value": round(seller_specialist_cost, 2),"formula": "Seller Revenue × Seller Specialist Split %"},
        "buyer_specialist_cost":   {"value": round(buyer_specialist_cost, 2), "formula": "Buyer Revenue × Buyer Specialist Split %"},
        "total_cost_of_sale":      {"value": round(total_cost_of_sale, 2),    "formula": "Paid to Brokerage + Referral Fees + Seller Specialist + Buyer Specialist"},
        "gross_profit":            {"value": round(gross_profit, 2),          "formula": "GCI Target − Total Cost of Sale"},
        "total_operating_expenses":{"value": round(total_op_exp, 2),          "formula": "Sum of all Operating Expense line items"},
        "net_income":              {"value": round(net_income, 2),            "formula": "Gross Profit − Total Operating Expenses"},
        "charity_amount":          {"value": round(charity_amt, 2),           "formula": "Net Income × Charity %"},
        "retirement_amount":       {"value": round(retirement_amt, 2),        "formula": "Net Income × Retirement %"},
        "taxable_income":          {"value": round(taxable_income, 2),        "formula": "Net Income − Charity − Retirement"},
        "tax_amount":              {"value": round(tax_amt, 2),               "formula": "Taxable Income × Income Tax %"},
        "take_home":               {"value": round(take_home, 2),             "formula": "Taxable Income − Tax Amount"},
        "monthly_life_budget":     {"value": round(monthly_budget, 2),        "formula": "Take Home ÷ 12"},
        "annual_personal_expenses":{"value": round(annual_personal, 2),       "formula": "Monthly Personal Subtotal × 12"},
        "annual_need_w_tax":       {"value": round(annual_need_w_tax, 2),     "formula": "(Annual Personal + Operating Expenses) × 1.30"},
        "survival_closings":       {"value": round(survival_closings, 2),     "formula": "Annual Need (incl. tax) ÷ Avg Net Commission per Close"},
        "closings_beyond_survival":{"value": round(closings_beyond_survival, 2), "formula": "Total Closings (planned) − Survival Closings"},
        "surplus":                 {"value": round(surplus, 2),               "formula": "Closings Beyond Survival × Avg Net Commission × (1 − Tax %)"},
    }


# ════════════════════════════════════════════════════════════
# Math: Lead Gen derivations
# ════════════════════════════════════════════════════════════

def calc_lead_gen(econ_calc: dict, lg: dict) -> dict:
    met_pct = float(lg.get("met_pct") or 0.8)
    cur_met = float(lg.get("current_met_db_size") or 0)
    cur_havent = float(lg.get("current_havent_met_db_size") or 0)

    total_sales = econ_calc["houses_sold_total"]["value"]
    sales_from_met = total_sales * met_pct
    sales_from_havent_met = total_sales - sales_from_met

    # MREA rule: 12 touches/yr × 2 contacts per sale → required Met DB = met sales × 12 / 2 = met sales × 6
    required_met_db = sales_from_met * 6
    # Haven't-Met rule: 50:1 ratio
    required_havent_met_db = sales_from_havent_met * 50

    return {
        "sales_from_met":          {"value": round(sales_from_met, 2),       "formula": "Total Sales × Met %"},
        "sales_from_havent_met":   {"value": round(sales_from_havent_met, 2),"formula": "Total Sales − Sales from Met"},
        "required_met_db":         {"value": round(required_met_db, 0),     "formula": "Met Sales × 12 ÷ 2  (12 touches/yr, 2 contacts per sale)"},
        "required_havent_met_db":  {"value": round(required_havent_met_db, 0), "formula": "Haven't-Met Sales × 50"},
        "met_gap":                 {"value": round(required_met_db - cur_met, 0), "formula": "Required Met DB − Current Met DB"},
        "havent_met_gap":          {"value": round(required_havent_met_db - cur_havent, 0), "formula": "Required Haven't-Met DB − Current Haven't-Met DB"},
    }


# ════════════════════════════════════════════════════════════
# Routes — Coaching Clients
# ════════════════════════════════════════════════════════════

@router.get("/clients")
def list_clients(request: Request):
    workspace_id = _ws(request)
    resp = _db("coaching_clients").select("*").order("created_at", desc=True).execute()
    rows = resp.data or []
    return [_enrich_client(r) for r in rows]


@router.get("/clients/{client_id}")
def get_client(client_id: int, request: Request):
    resp = _db("coaching_clients").select("*").eq("id", client_id).execute()
    if not resp.data:
        raise HTTPException(404, "Coaching client not found")
    return _enrich_client(resp.data[0])


@router.post("/clients")
def create_client(payload: CoachingClientIn, request: Request):
    workspace_id = _ws(request)
    lead_id = payload.lead_id

    # New-contact path — create the lead first
    if not lead_id and payload.new_contact:
        nc = payload.new_contact
        first = (nc.get("first_name") or "").strip()
        last = (nc.get("last_name") or "").strip()
        full_name = (nc.get("name") or f"{first} {last}").strip()
        email = (nc.get("email") or "").strip()
        if not full_name or not email:
            raise HTTPException(400, "New contact requires name and email")
        # Reuse if a lead with this email already exists in the workspace
        existing = _db("leads").select("id").eq("email", email).execute()
        if existing.data:
            lead_id = existing.data[0]["id"]
        else:
            lead_row = {
                "first_name": first or full_name.split(" ")[0],
                "last_name": last or " ".join(full_name.split(" ")[1:]),
                "name": full_name,
                "email": email,
                "phone": nc.get("phone"),
                "current_brokerage": nc.get("current_brokerage"),
                "stage": "coaching_client",
                "source": "Coaching",
            }
            ins = _db("leads").insert(lead_row).execute()
            lead_id = ins.data[0]["id"]

    if not lead_id:
        raise HTTPException(400, "Either lead_id or new_contact is required")

    # Refuse to duplicate
    existing_cc = _supabase.table("coaching_clients").select("id").eq("lead_id", lead_id).execute()
    if existing_cc.data:
        raise HTTPException(409, "This contact is already a coaching client")

    row = {
        "workspace_id": workspace_id,
        "lead_id": lead_id,
        "brokerage": payload.brokerage,
        "lpt_comp_plan": payload.lpt_comp_plan,
        "license_date": payload.license_date,
        "market_city": payload.market_city,
        "market_state": payload.market_state,
        "avg_sale_price": payload.avg_sale_price,
        "avg_commission_rate": _normalize_pct(payload.avg_commission_rate),
        "call_cadence": payload.call_cadence or "WEEKLY",
        "coaching_start_date": payload.coaching_start_date or date.today().isoformat(),
        "notes": payload.notes,
    }
    row = {k: v for k, v in row.items() if v is not None}
    ins = _supabase.table("coaching_clients").insert(row).execute()
    client = ins.data[0]

    # Auto-create the current-year business plan + child models so the editor has data to render
    _ensure_business_plan(client["id"], datetime.utcnow().year, workspace_id)

    # Auto-provision the portal + send invite email if requested
    portal_result = None
    if payload.send_portal_invite:
        try:
            portal_result = provision_portal(client["id"], ProvisionPortalIn(send_email=True), request)
        except Exception as e:
            portal_result = {"error": str(e)}

    enriched = _enrich_client(client)
    if portal_result is not None:
        enriched["portal_provision"] = portal_result
    return enriched


@router.patch("/clients/{client_id}")
def update_client(client_id: int, payload: CoachingClientUpdate, request: Request):
    data = {k: v for k, v in payload.dict().items() if v is not None}
    if "avg_commission_rate" in data:
        data["avg_commission_rate"] = _normalize_pct(data["avg_commission_rate"])
    if not data:
        raise HTTPException(400, "No fields to update")
    resp = _supabase.table("coaching_clients").update(data).eq("id", client_id).execute()
    if not resp.data:
        raise HTTPException(404, "Coaching client not found")
    return _enrich_client(resp.data[0])


@router.delete("/clients/{client_id}")
def delete_client(client_id: int, request: Request):
    _supabase.table("coaching_clients").delete().eq("id", client_id).execute()
    return {"ok": True}


# ════════════════════════════════════════════════════════════
# Routes — Business Plan + child models
# ════════════════════════════════════════════════════════════

@router.get("/clients/{client_id}/plan")
def get_plan(client_id: int, request: Request, year: Optional[int] = None):
    workspace_id = _ws(request)
    yr = year or datetime.utcnow().year
    bundle = _ensure_business_plan(client_id, yr, workspace_id)
    return bundle


@router.patch("/clients/{client_id}/plan")
def update_plan(client_id: int, payload: BusinessPlanUpdate, request: Request, year: Optional[int] = None):
    yr = year or datetime.utcnow().year
    plan_resp = _supabase.table("business_plans").select("id").eq("coaching_client_id", client_id).eq("year", yr).execute()
    if not plan_resp.data:
        raise HTTPException(404, f"No plan for year {yr}")
    plan_id = plan_resp.data[0]["id"]
    data = {k: v for k, v in payload.dict().items() if v is not None}
    if not data:
        raise HTTPException(400, "No fields to update")
    upd = _supabase.table("business_plans").update(data).eq("id", plan_id).execute()
    return upd.data[0]


@router.patch("/clients/{client_id}/budget-model")
def update_budget(client_id: int, payload: BudgetModelUpdate, request: Request, year: Optional[int] = None):
    yr = year or datetime.utcnow().year
    plan_resp = _supabase.table("business_plans").select("id").eq("coaching_client_id", client_id).eq("year", yr).execute()
    if not plan_resp.data:
        raise HTTPException(404, f"No plan for year {yr}")
    plan_id = plan_resp.data[0]["id"]
    data = payload.dict(exclude_unset=True)
    # Normalize percentages
    for k in ("referrals_split_pct", "seller_specialist_split_pct", "buyer_specialist_split_pct",
              "charity_pct", "retirement_pct", "income_tax_pct"):
        if k in data and data[k] is not None:
            data[k] = _normalize_pct(data[k])
    if not data:
        raise HTTPException(400, "No fields to update")
    upd = _supabase.table("budget_models").update(data).eq("business_plan_id", plan_id).execute()
    if not upd.data:
        raise HTTPException(404, "Budget model not found")
    return upd.data[0]


@router.patch("/clients/{client_id}/economic-model")
def update_economic(client_id: int, payload: EconomicModelUpdate, request: Request, year: Optional[int] = None):
    yr = year or datetime.utcnow().year
    plan_resp = _supabase.table("business_plans").select("id").eq("coaching_client_id", client_id).eq("year", yr).execute()
    if not plan_resp.data:
        raise HTTPException(404, f"No plan for year {yr}")
    plan_id = plan_resp.data[0]["id"]
    data = payload.dict(exclude_unset=True)
    for k in ("seller_pct", "commission_rate", "listings_close_pct", "buyers_close_pct",
              "listing_appt_to_list_pct", "buyer_appt_to_work_pct"):
        if k in data and data[k] is not None:
            data[k] = _normalize_pct(data[k])
    if not data:
        raise HTTPException(400, "No fields to update")
    upd = _supabase.table("economic_models").update(data).eq("business_plan_id", plan_id).execute()
    if not upd.data:
        raise HTTPException(404, "Economic model not found")
    return upd.data[0]


# ════════════════════════════════════════════════════════════
# Routes — Computed numbers (auditable)
# ════════════════════════════════════════════════════════════

@router.get("/clients/{client_id}/computed")
def get_computed(client_id: int, request: Request, year: Optional[int] = None):
    workspace_id = _ws(request)
    yr = year or datetime.utcnow().year
    bundle = _ensure_business_plan(client_id, yr, workspace_id)
    plan = bundle["plan"]
    bm = bundle["budget_model"]
    em = bundle["economic_model"]
    lg = bundle["lead_gen_model"]
    gci = float(plan.get("gci_target") or 0)

    economic = calc_economic(gci, em)
    budget = calc_budget(gci, bm, economic)
    leadgen = calc_lead_gen(economic, lg)

    return {
        "year": yr,
        "gci_target": gci,
        "economic": economic,
        "budget": budget,
        "lead_gen": leadgen,
    }


# ════════════════════════════════════════════════════════════
# Routes — Lead picker (for the Invite modal)
# ════════════════════════════════════════════════════════════

@router.get("/lead-search")
def search_leads(request: Request, q: Optional[str] = None, limit: int = 15):
    """Autocomplete for the Invite modal — returns candidate leads not yet flagged as coaching clients."""
    workspace_id = _ws(request)
    q_str = (q or "").strip()
    qb = _db("leads").select("id,first_name,last_name,name,email,phone,current_brokerage")
    if q_str:
        # Supabase OR filter syntax
        like = f"%{q_str}%"
        qb = qb.or_(f"name.ilike.{like},email.ilike.{like},first_name.ilike.{like},last_name.ilike.{like}")
    leads = qb.limit(limit).execute().data or []
    # Filter out leads already linked to a coaching_client
    if leads:
        ids = [l["id"] for l in leads]
        existing = _supabase.table("coaching_clients").select("lead_id").in_("lead_id", ids).execute().data or []
        existing_set = {r["lead_id"] for r in existing}
        leads = [l for l in leads if l["id"] not in existing_set]
    return leads


# ════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════
# SESSION 2 — Coaching Calls / Action Items / Activity / Pipeline
# ════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════

# ─── Pydantic ───

class PipelineEntryIn(BaseModel):
    entry_type: Optional[str] = None  # LISTING | BUYER (only required on create)
    appointment_date: Optional[str] = None
    address: Optional[str] = None
    contact_name: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_email: Optional[str] = None
    next_step: Optional[str] = None
    notes: Optional[str] = None
    rating: Optional[int] = None
    # Phase 15.5c — transaction lifecycle
    status: Optional[str] = None  # PRE_SIGNED|ACTIVE|PENDING|CLOSED|EXPIRED|WITHDRAWN|CANCELLED
    list_date: Optional[str] = None
    expiration_date: Optional[str] = None
    expected_close_date: Optional[str] = None
    mls_number: Optional[str] = None
    closed: Optional[bool] = None
    closed_date: Optional[str] = None
    closing_price: Optional[float] = None
    gross_commission: Optional[float] = None
    net_commission: Optional[float] = None


class ActivityLogIn(BaseModel):
    log_date: Optional[str] = None
    # Legacy aggregate fields (kept for backward compat; new UI uses the granular ones below)
    contacts_made: Optional[int] = 0
    appts_set: Optional[int] = 0
    appts_held: Optional[int] = 0
    hours_prospected: Optional[float] = 0
    # Phase 15.5 — CTE-style granular daily lead gen entry
    dials: Optional[int] = 0
    nurtures: Optional[int] = 0
    listing_appts_set: Optional[int] = 0
    listing_appts_held: Optional[int] = 0
    listings_signed: Optional[int] = 0
    buyer_appts_set: Optional[int] = 0
    buyer_appts_held: Optional[int] = 0
    buyers_signed: Optional[int] = 0
    showings: Optional[int] = 0
    open_houses_held: Optional[int] = 0
    wins: Optional[str] = None
    notes: Optional[str] = None


class CoachingCallIn(BaseModel):
    scheduled_at: Optional[str] = None  # ISO datetime
    call_type: Optional[str] = "WEEKLY"
    in_call_notes: Optional[str] = None
    status: Optional[str] = None
    completed_at: Optional[str] = None


class ActionItemIn(BaseModel):
    text: str
    measurement: Optional[str] = None
    due_date: Optional[str] = None
    owner: Optional[str] = "AGENT"
    tag: Optional[str] = None
    status: Optional[str] = "OPEN"
    source_call_id: Optional[int] = None


class ActionItemUpdate(BaseModel):
    text: Optional[str] = None
    measurement: Optional[str] = None
    due_date: Optional[str] = None
    owner: Optional[str] = None
    tag: Optional[str] = None
    status: Optional[str] = None


# ─── Pipeline entries ───

@router.get("/clients/{client_id}/pipeline")
def list_pipeline(client_id: int, request: Request, entry_type: Optional[str] = None):
    qb = _supabase.table("pipeline_entries").select("*").eq("coaching_client_id", client_id)
    if entry_type:
        qb = qb.eq("entry_type", entry_type)
    res = qb.order("rating", desc=True).order("appointment_date", desc=True).execute()
    return res.data or []


@router.post("/clients/{client_id}/pipeline")
def create_pipeline(client_id: int, payload: PipelineEntryIn, request: Request):
    workspace_id = _ws(request)
    if payload.entry_type not in ("LISTING", "BUYER"):
        raise HTTPException(400, "entry_type must be LISTING or BUYER")
    if payload.rating is not None and not (1 <= payload.rating <= 10):
        raise HTTPException(400, "rating must be 1–10")
    row = {"workspace_id": workspace_id, "coaching_client_id": client_id, **payload.dict(exclude_unset=True)}
    res = _supabase.table("pipeline_entries").insert(row).execute()
    return res.data[0]


@router.patch("/pipeline/{entry_id}")
def update_pipeline(entry_id: int, payload: PipelineEntryIn, request: Request):
    data = payload.dict(exclude_unset=True)
    if not data:
        raise HTTPException(400, "No fields to update")
    res = _supabase.table("pipeline_entries").update(data).eq("id", entry_id).execute()
    if not res.data:
        raise HTTPException(404, "Pipeline entry not found")
    return res.data[0]


@router.delete("/pipeline/{entry_id}")
def delete_pipeline(entry_id: int, request: Request):
    _supabase.table("pipeline_entries").delete().eq("id", entry_id).execute()
    return {"ok": True}


# ─── Activity logs (daily) ───

@router.get("/clients/{client_id}/activity")
def list_activity(client_id: int, request: Request, days: int = 30):
    res = _supabase.table("coaching_activity_logs").select("*").eq("coaching_client_id", client_id).order("log_date", desc=True).limit(max(1, min(days, 365))).execute()
    return res.data or []


@router.post("/clients/{client_id}/activity")
def upsert_activity(client_id: int, payload: ActivityLogIn, request: Request):
    """Upsert by (client_id, log_date) so re-saving today's entry doesn't dupe."""
    log_date = payload.log_date or date.today().isoformat()
    row = {
        "coaching_client_id": client_id,
        "log_date": log_date,
        **{k: v for k, v in payload.dict(exclude_unset=True).items() if k != "log_date"},
    }
    res = _supabase.table("coaching_activity_logs").upsert(row, on_conflict="coaching_client_id,log_date").execute()
    return res.data[0] if res.data else row


def _activity_streak(client_id: int, target_contacts: int = 1) -> int:
    """Count consecutive days back from today where contacts_made >= target."""
    res = _supabase.table("coaching_activity_logs").select("log_date,contacts_made").eq("coaching_client_id", client_id).order("log_date", desc=True).limit(120).execute()
    rows = res.data or []
    if not rows:
        return 0
    today = date.today()
    streak = 0
    expected = today
    for r in rows:
        ld = r.get("log_date")
        if not ld:
            continue
        try:
            d = datetime.strptime(ld, "%Y-%m-%d").date()
        except Exception:
            continue
        if d == expected and (r.get("contacts_made") or 0) >= target_contacts:
            streak += 1
            expected = expected.replace(day=expected.day) - timedelta_days(1)
        else:
            break
    return streak


# datetime.timedelta; declare a small helper to avoid importing timedelta separately.
def timedelta_days(n):
    from datetime import timedelta
    return timedelta(days=n)


# ─── Coaching calls ───

@router.get("/clients/{client_id}/calls")
def list_calls(client_id: int, request: Request, limit: int = 50):
    res = _db("coaching_calls").select("*").eq("coaching_client_id", client_id).order("scheduled_at", desc=True).limit(limit).execute()
    return res.data or []


@router.get("/calls/{call_id}")
def get_call(call_id: int, request: Request):
    res = _db("coaching_calls").select("*").eq("id", call_id).execute()
    if not res.data:
        raise HTTPException(404, "Call not found")
    return res.data[0]


@router.post("/clients/{client_id}/calls")
def create_call(client_id: int, payload: CoachingCallIn, request: Request):
    """Schedule a new call. Auto-snapshots the pre-call brief at scheduled time."""
    workspace_id = _ws(request)
    # Find prior call for commitment-keep score chaining
    prior = _supabase.table("coaching_calls").select("id").eq("coaching_client_id", client_id).order("scheduled_at", desc=True).limit(1).execute()
    prior_id = prior.data[0]["id"] if prior.data else None

    row = {
        "workspace_id": workspace_id,
        "coaching_client_id": client_id,
        "scheduled_at": payload.scheduled_at or datetime.utcnow().isoformat(),
        "call_type": payload.call_type or "WEEKLY",
        "status": "SCHEDULED",
        "prior_call_id": prior_id,
    }
    res = _supabase.table("coaching_calls").insert(row).execute()
    call = res.data[0]
    # Pre-snapshot the brief on the call so it's preserved even if numbers change later
    try:
        brief = _build_brief(client_id, call["id"], workspace_id)
        _supabase.table("coaching_calls").update({"pre_call_brief": brief}).eq("id", call["id"]).execute()
        call["pre_call_brief"] = brief
    except Exception:
        pass
    return call


@router.patch("/calls/{call_id}")
def update_call(call_id: int, payload: CoachingCallIn, request: Request):
    data = payload.dict(exclude_unset=True)
    if not data:
        raise HTTPException(400, "No fields to update")
    if data.get("status") == "COMPLETED" and not data.get("completed_at"):
        data["completed_at"] = datetime.utcnow().isoformat()
    res = _supabase.table("coaching_calls").update(data).eq("id", call_id).execute()
    if not res.data:
        raise HTTPException(404, "Call not found")
    return res.data[0]


@router.delete("/calls/{call_id}")
def delete_call(call_id: int, request: Request):
    _supabase.table("coaching_calls").delete().eq("id", call_id).execute()
    return {"ok": True}


@router.post("/calls/{call_id}/refresh-brief")
def refresh_brief(call_id: int, request: Request):
    """Re-build the pre-call brief from current data (useful right before the call starts)."""
    workspace_id = _ws(request)
    call_resp = _supabase.table("coaching_calls").select("coaching_client_id").eq("id", call_id).single().execute()
    if not call_resp.data:
        raise HTTPException(404, "Call not found")
    client_id = call_resp.data["coaching_client_id"]
    brief = _build_brief(client_id, call_id, workspace_id)
    _supabase.table("coaching_calls").update({"pre_call_brief": brief}).eq("id", call_id).execute()
    return brief


@router.post("/calls/{call_id}/complete")
def complete_call(call_id: int, request: Request):
    """Mark the call complete + compute commitment-keep score by counting completion of action items
    that were set on the PRIOR call."""
    call = _supabase.table("coaching_calls").select("*").eq("id", call_id).single().execute().data
    if not call:
        raise HTTPException(404, "Call not found")
    prior_id = call.get("prior_call_id")
    score = None
    if prior_id:
        items = _supabase.table("coaching_action_items").select("status").eq("source_call_id", prior_id).execute().data or []
        if items:
            completed = sum(1 for it in items if (it.get("status") or "").upper() == "COMPLETED")
            score = round((completed / len(items)) * 100.0, 1)
    update = {
        "status": "COMPLETED",
        "completed_at": datetime.utcnow().isoformat(),
    }
    if score is not None:
        update["commitment_keep_score"] = score
    res = _supabase.table("coaching_calls").update(update).eq("id", call_id).execute()
    return res.data[0]


# ─── Action items ───

@router.get("/clients/{client_id}/action-items")
def list_action_items(client_id: int, request: Request, status: Optional[str] = None, source_call_id: Optional[int] = None):
    qb = _db("coaching_action_items").select("*").eq("coaching_client_id", client_id)
    if status:
        qb = qb.eq("status", status)
    if source_call_id is not None:
        qb = qb.eq("source_call_id", source_call_id)
    res = qb.order("due_date", desc=False).order("created_at", desc=True).execute()
    return res.data or []


@router.post("/clients/{client_id}/action-items")
def create_action_item(client_id: int, payload: ActionItemIn, request: Request):
    workspace_id = _ws(request)
    row = {
        "workspace_id": workspace_id,
        "coaching_client_id": client_id,
        **payload.dict(exclude_unset=True),
    }
    res = _supabase.table("coaching_action_items").insert(row).execute()
    return res.data[0]


@router.patch("/action-items/{item_id}")
def update_action_item(item_id: int, payload: ActionItemUpdate, request: Request):
    data = payload.dict(exclude_unset=True)
    if not data:
        raise HTTPException(400, "No fields to update")
    if data.get("status") == "COMPLETED":
        data["completed_at"] = datetime.utcnow().isoformat()
    res = _supabase.table("coaching_action_items").update(data).eq("id", item_id).execute()
    if not res.data:
        raise HTTPException(404, "Action item not found")
    return res.data[0]


@router.delete("/action-items/{item_id}")
def delete_action_item(item_id: int, request: Request):
    _supabase.table("coaching_action_items").delete().eq("id", item_id).execute()
    return {"ok": True}


# ─── Pre-call brief generator ───
# Pulls from plan, pipeline, activity logs, last call's action items.
# Returns structured JSON: every section is a list of named metrics with status flags.

def _build_brief(client_id: int, call_id: int, workspace_id: int) -> dict:
    today = date.today()
    yr = today.year

    # Client + plan + computed
    client = _supabase.table("coaching_clients").select("*").eq("id", client_id).single().execute().data or {}
    bundle = _ensure_business_plan(client_id, yr, workspace_id)
    plan = bundle["plan"]
    em = bundle["economic_model"]
    bm = bundle["budget_model"]
    gci_target = float(plan.get("gci_target") or 0)
    economic = calc_economic(gci_target, em)

    # GCI YTD = sum of gross_commission on closed pipeline entries this year
    closed_resp = _supabase.table("pipeline_entries").select("gross_commission,closed_date,entry_type").eq("coaching_client_id", client_id).eq("closed", True).execute()
    gci_ytd = 0
    listings_closed = 0
    buyers_closed = 0
    for r in (closed_resp.data or []):
        cd = r.get("closed_date")
        if not cd or not cd.startswith(str(yr)):
            continue
        gci_ytd += float(r.get("gross_commission") or 0)
        if r.get("entry_type") == "LISTING":
            listings_closed += 1
        elif r.get("entry_type") == "BUYER":
            buyers_closed += 1

    # Pace = (YTD ÷ target) vs (days elapsed ÷ 365). Negative = behind; zero = on pace.
    days_elapsed = (today - date(yr, 1, 1)).days + 1
    expected_pct = days_elapsed / 365.0
    actual_pct = (gci_ytd / gci_target) if gci_target > 0 else 0
    pace_gap_pct = round((actual_pct - expected_pct) * 100, 1)
    if pace_gap_pct >= 0:
        pace_status = "ahead"
    elif pace_gap_pct >= -10:
        pace_status = "on-pace"
    else:
        pace_status = "behind"

    # Pipeline by rating
    open_pipe = _supabase.table("pipeline_entries").select("entry_type,rating").eq("coaching_client_id", client_id).eq("closed", False).execute().data or []
    def count_at(et, r):
        return sum(1 for p in open_pipe if p.get("entry_type") == et and p.get("rating") == r)
    pipeline = {
        "listings": {
            "10s": count_at("LISTING", 10), "9s": count_at("LISTING", 9), "8s": count_at("LISTING", 8),
            "7s": count_at("LISTING", 7), "6s": count_at("LISTING", 6),
            "cold": sum(1 for p in open_pipe if p.get("entry_type") == "LISTING" and (p.get("rating") or 0) <= 5),
            "total": sum(1 for p in open_pipe if p.get("entry_type") == "LISTING"),
        },
        "buyers": {
            "10s": count_at("BUYER", 10), "9s": count_at("BUYER", 9), "8s": count_at("BUYER", 8),
            "7s": count_at("BUYER", 7), "6s": count_at("BUYER", 6),
            "cold": sum(1 for p in open_pipe if p.get("entry_type") == "BUYER" and (p.get("rating") or 0) <= 5),
            "total": sum(1 for p in open_pipe if p.get("entry_type") == "BUYER"),
        },
    }

    # Activity (last 14 days)
    from datetime import timedelta as _td
    cutoff = (today - _td(days=14)).isoformat()
    act_resp = _supabase.table("coaching_activity_logs").select("*").eq("coaching_client_id", client_id).gte("log_date", cutoff).execute()
    acts = act_resp.data or []
    contacts_14d = sum((a.get("contacts_made") or 0) for a in acts)
    appts_held_14d = sum((a.get("appts_held") or 0) for a in acts)
    hours_14d = sum(float(a.get("hours_prospected") or 0) for a in acts)
    days_logged_14d = len(acts)

    # Streak — consecutive days from today with contacts_made >= 1
    streak = _activity_streak(client_id, target_contacts=1)

    # Last call's action items + commitment-keep
    call_row = _supabase.table("coaching_calls").select("prior_call_id").eq("id", call_id).single().execute().data or {}
    prior_id = call_row.get("prior_call_id")
    prior_items = []
    commit_score = None
    if prior_id:
        prior_items = _supabase.table("coaching_action_items").select("text,measurement,status,due_date").eq("source_call_id", prior_id).execute().data or []
        if prior_items:
            done = sum(1 for it in prior_items if (it.get("status") or "").upper() == "COMPLETED")
            commit_score = round((done / len(prior_items)) * 100.0, 1)

    # Suggested talking points — heuristic
    talking = []
    if pace_status == "behind":
        talking.append(f"GCI pace is {abs(pace_gap_pct)}% behind. Gap = ${round((expected_pct * gci_target) - gci_ytd):,}.")
    if pipeline["listings"]["10s"] + pipeline["listings"]["9s"] == 0 and pipeline["listings"]["total"] > 0:
        talking.append("No 10s or 9s in listing pipeline — no listings expected to close in next 60 days.")
    if pipeline["listings"]["total"] == 0:
        talking.append("Listing pipeline is EMPTY — lead-gen problem, not skill problem.")
    if commit_score is not None and commit_score < 70:
        talking.append(f"Commitment-keep score = {commit_score}% from last call. Below 70% — coachability check.")
    if days_logged_14d < 7:
        talking.append(f"Only {days_logged_14d} of last 14 days logged activity. Daily discipline gap.")
    if contacts_14d < 50 and days_logged_14d > 0:
        talking.append(f"Only {contacts_14d} contacts made in 14 days. Way under prospecting target.")

    return {
        "generated_at": datetime.utcnow().isoformat(),
        "client_id": client_id,
        "year": yr,

        "header": {
            "name": (client.get("notes") or "")[:0] or None,  # placeholder, UI renders from /clients/{id}
            "comp_plan": client.get("lpt_comp_plan"),
            "call_cadence": client.get("call_cadence"),
        },

        "pace": {
            "gci_target": gci_target,
            "gci_ytd": round(gci_ytd, 2),
            "expected_pct": round(expected_pct * 100, 1),
            "actual_pct": round(actual_pct * 100, 1),
            "pace_gap_pct": pace_gap_pct,
            "status": pace_status,            # behind | on-pace | ahead
            "days_elapsed": days_elapsed,
        },

        "big_rocks": {
            "listings_taken_target": economic["listings_taken"]["value"],
            "listings_closed_ytd":   listings_closed,
            "buyers_closed_ytd":     buyers_closed,
            "houses_sold_target":    economic["houses_sold_total"]["value"],
            "listing_appts_target":  economic["listing_appts_annual"]["value"],
            "buyer_consults_target": economic["buyer_consults_annual"]["value"],
        },

        "pipeline": pipeline,

        "activity_14d": {
            "contacts_made":   contacts_14d,
            "appts_held":      appts_held_14d,
            "hours_prospected": round(hours_14d, 1),
            "days_logged":     days_logged_14d,
            "streak_days":     streak,
        },

        "last_call_action_items": prior_items,
        "commitment_keep_score": commit_score,

        "talking_points": talking,
    }


@router.get("/clients/{client_id}/brief-preview")
def preview_brief(client_id: int, request: Request):
    """Build an ad-hoc brief without creating a call. Useful for the dashboard."""
    workspace_id = _ws(request)
    # Use a fake call_id of 0 so the prior-call lookup just returns nothing
    today = date.today()
    yr = today.year
    bundle = _ensure_business_plan(client_id, yr, workspace_id)
    # Build with no call_id (no prior commitment chain)
    fake_call = {"prior_call_id": None}
    # We want the full brief output minus prior items — easiest path is to insert a phantom into _build_brief.
    # Instead just call with the most recent call_id if any, else build manually.
    last_call = _supabase.table("coaching_calls").select("id").eq("coaching_client_id", client_id).order("scheduled_at", desc=True).limit(1).execute()
    if last_call.data:
        return _build_brief(client_id, last_call.data[0]["id"], workspace_id)
    # No prior call — synthesize by calling _build_brief with -1 and catching the empty prior_id branch
    # (the function gracefully handles missing prior call rows)
    # Insert a temporary stub that won't be saved
    return _build_brief_no_call(client_id, workspace_id)


def _build_brief_no_call(client_id: int, workspace_id: int) -> dict:
    """Same as _build_brief but with no prior-call chain (used for first-ever brief)."""
    today = date.today()
    yr = today.year
    client = _supabase.table("coaching_clients").select("*").eq("id", client_id).single().execute().data or {}
    bundle = _ensure_business_plan(client_id, yr, workspace_id)
    plan = bundle["plan"]
    em = bundle["economic_model"]
    gci_target = float(plan.get("gci_target") or 0)
    economic = calc_economic(gci_target, em)

    closed_resp = _supabase.table("pipeline_entries").select("gross_commission,closed_date,entry_type").eq("coaching_client_id", client_id).eq("closed", True).execute()
    gci_ytd, listings_closed, buyers_closed = 0, 0, 0
    for r in (closed_resp.data or []):
        cd = r.get("closed_date")
        if not cd or not cd.startswith(str(yr)):
            continue
        gci_ytd += float(r.get("gross_commission") or 0)
        if r.get("entry_type") == "LISTING": listings_closed += 1
        elif r.get("entry_type") == "BUYER": buyers_closed += 1

    days_elapsed = (today - date(yr, 1, 1)).days + 1
    expected_pct = days_elapsed / 365.0
    actual_pct = (gci_ytd / gci_target) if gci_target > 0 else 0
    pace_gap_pct = round((actual_pct - expected_pct) * 100, 1)
    pace_status = "ahead" if pace_gap_pct >= 0 else ("on-pace" if pace_gap_pct >= -10 else "behind")

    open_pipe = _supabase.table("pipeline_entries").select("entry_type,rating").eq("coaching_client_id", client_id).eq("closed", False).execute().data or []
    def cnt(et, r):
        return sum(1 for p in open_pipe if p.get("entry_type") == et and p.get("rating") == r)
    pipeline = {
        "listings": {"10s": cnt("LISTING",10), "9s": cnt("LISTING",9), "8s": cnt("LISTING",8),
                     "7s": cnt("LISTING",7), "6s": cnt("LISTING",6),
                     "cold": sum(1 for p in open_pipe if p.get("entry_type")=="LISTING" and (p.get("rating") or 0)<=5),
                     "total": sum(1 for p in open_pipe if p.get("entry_type")=="LISTING")},
        "buyers":   {"10s": cnt("BUYER",10), "9s": cnt("BUYER",9), "8s": cnt("BUYER",8),
                     "7s": cnt("BUYER",7), "6s": cnt("BUYER",6),
                     "cold": sum(1 for p in open_pipe if p.get("entry_type")=="BUYER" and (p.get("rating") or 0)<=5),
                     "total": sum(1 for p in open_pipe if p.get("entry_type")=="BUYER")},
    }

    from datetime import timedelta as _td
    cutoff = (today - _td(days=14)).isoformat()
    acts = (_supabase.table("coaching_activity_logs").select("*").eq("coaching_client_id", client_id).gte("log_date", cutoff).execute().data) or []
    streak = _activity_streak(client_id, target_contacts=1)

    talking = []
    if pace_status == "behind":
        talking.append(f"GCI pace is {abs(pace_gap_pct)}% behind. Gap = ${round((expected_pct * gci_target) - gci_ytd):,}.")
    if pipeline["listings"]["total"] == 0:
        talking.append("Listing pipeline is EMPTY — lead-gen problem.")

    return {
        "generated_at": datetime.utcnow().isoformat(),
        "client_id": client_id,
        "year": yr,
        "pace": {
            "gci_target": gci_target, "gci_ytd": round(gci_ytd, 2),
            "expected_pct": round(expected_pct*100,1), "actual_pct": round(actual_pct*100,1),
            "pace_gap_pct": pace_gap_pct, "status": pace_status, "days_elapsed": days_elapsed,
        },
        "big_rocks": {
            "listings_taken_target": economic["listings_taken"]["value"],
            "listings_closed_ytd": listings_closed,
            "buyers_closed_ytd": buyers_closed,
            "houses_sold_target": economic["houses_sold_total"]["value"],
            "listing_appts_target": economic["listing_appts_annual"]["value"],
            "buyer_consults_target": economic["buyer_consults_annual"]["value"],
        },
        "pipeline": pipeline,
        "activity_14d": {
            "contacts_made": sum((a.get("contacts_made") or 0) for a in acts),
            "appts_held": sum((a.get("appts_held") or 0) for a in acts),
            "hours_prospected": round(sum(float(a.get("hours_prospected") or 0) for a in acts), 1),
            "days_logged": len(acts),
            "streak_days": streak,
        },
        "last_call_action_items": [],
        "commitment_keep_score": None,
        "talking_points": talking,
    }


# ════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════
# SESSION 3 — Agent Portal /coaching + Coaching Dashboard +
#             GPS / 4-1-1 + HybridShare
# ════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════

# ─── Portal provisioning (creates a user the agent logs in with) ───

class ProvisionPortalIn(BaseModel):
    send_email: Optional[bool] = True


class ResendInviteIn(BaseModel):
    new_email: Optional[str] = None     # if set, updates the lead + user records first
    remove_suppression: Optional[bool] = True


@router.post("/clients/{client_id}/resend-invite")
def resend_invite(client_id: int, payload: ResendInviteIn, request: Request):
    """Re-issue the portal invite — useful when the original bounced or the agent typed the wrong email.
    Optionally updates the lead + user email first, removes the address from the suppression list, and
    regenerates a fresh temp password."""
    import secrets as _secrets
    from auth import hash_password as _hp

    cc = _supabase.table("coaching_clients").select("*").eq("id", client_id).single().execute().data
    if not cc:
        raise HTTPException(404, "Coaching client not found")

    # Step 1: optional email update
    new_email = (payload.new_email or "").strip().lower()
    if new_email:
        if not is_valid_email(new_email):
            raise HTTPException(400, "Invalid email address")
        # Update the lead
        _supabase.table("leads").update({"email": new_email}).eq("id", cc["lead_id"]).execute()
        # Update the user if one exists
        if cc.get("user_id"):
            _supabase.table("users").update({"email": new_email}).eq("id", cc["user_id"]).execute()

    # Resolve the current email (after any update)
    lead = _supabase.table("leads").select("*").eq("id", cc["lead_id"]).single().execute().data or {}
    email = (lead.get("email") or "").strip().lower()
    if not email or not is_valid_email(email):
        raise HTTPException(400, "Lead has no valid email — provide new_email")
    name = lead.get("name") or ((lead.get("first_name") or "") + " " + (lead.get("last_name") or "")).strip()

    # Step 2: remove suppression if requested (so Resend's webhook block doesn't kill the send)
    if payload.remove_suppression:
        try:
            _supabase.table("email_suppressions").delete().eq("email", email).execute()
        except Exception:
            pass

    # Step 3: ensure user exists (provisions if missing)
    user_id = cc.get("user_id")
    if not user_id:
        # No user yet — fall back to provision-portal logic. Reuse if email already a user.
        existing = _supabase.table("users").select("id").eq("email", email).execute().data or []
        if existing:
            user_id = existing[0]["id"]
            _supabase.table("coaching_clients").update({"user_id": user_id}).eq("id", client_id).execute()
        else:
            # Create user + workspace fresh (mirrors provision_portal)
            initials = "".join([p[0] for p in name.split()[:2] if p]).upper() or "AG"
            ins_user = _supabase.table("users").insert({
                "email": email, "password_hash": "", "name": name, "role": "agent", "avatar_initials": initials[:2],
            }).execute()
            user_id = ins_user.data[0]["id"]
            ws_ins = _supabase.table("workspaces").insert({
                "owner_user_id": user_id, "name": name + "'s Workspace", "plan": "basic",
                "settings": {"coaching_only": True},
            }).execute()
            _supabase.table("users").update({"workspace_id": ws_ins.data[0]["id"]}).eq("id", user_id).execute()
            _supabase.table("coaching_clients").update({"user_id": user_id}).eq("id", client_id).execute()

    # Step 4: regenerate temp password and update the user
    temp_pwd = _secrets.token_urlsafe(8)
    _supabase.table("users").update({"password_hash": _hp(temp_pwd)}).eq("id", user_id).execute()

    # Step 5: send invite email
    sent_ok = False
    send_error = None
    try:
        from main import send_email as _send, load_settings as _load_settings
        settings = _load_settings()
        smtp_cfg = settings.get("smtp") or {}
        html = f"""
        <div style='font-family:sans-serif;max-width:560px;margin:0 auto;padding:20px'>
          <h2 style='color:#6c63ff'>Your TPL Coaching Portal is Ready</h2>
          <p>Hey {name.split(' ')[0]} —</p>
          <p>Log in here to view your business plan, log daily activity, track your pipeline, and see prep for our calls.</p>
          <p style='background:#f5f5fa;border-radius:8px;padding:14px;font-family:monospace;font-size:13px'>
            <strong>Login:</strong> https://portal.tplcollective.ai<br>
            <strong>Email:</strong> {email}<br>
            <strong>Temporary password:</strong> {temp_pwd}
          </p>
          <p>Change your password after first login. See you on our next call.</p>
          <p>— Joe</p>
        </div>
        """
        ok, err = _send(smtp_cfg, email, "Your TPL Coaching Portal is Ready", html, from_address="Joe DeSane <joe@tplcollective.co>", campaign="coaching-invite")
        sent_ok = ok
        if not ok: send_error = err
    except Exception as e:
        send_error = str(e)

    return {
        "ok": True,
        "email": email,
        "user_id": user_id,
        "temp_password": temp_pwd,
        "email_sent": sent_ok,
        "send_error": send_error,
    }


@router.get("/clients/{client_id}/invite-status")
def invite_status(client_id: int, request: Request):
    """Returns the most recent invite-email status for this client's email + suppression status."""
    cc = _supabase.table("coaching_clients").select("user_id,lead_id").eq("id", client_id).single().execute().data
    if not cc:
        raise HTTPException(404, "Coaching client not found")
    lead = _supabase.table("leads").select("email").eq("id", cc["lead_id"]).single().execute().data or {}
    email = (lead.get("email") or "").lower()
    out = {"email": email, "user_provisioned": bool(cc.get("user_id")), "last_invite": None, "is_suppressed": False, "suppression_reason": None}
    if not email:
        return out
    last = _supabase.table("email_send_log").select("status,error,created_at,subject").eq("to_address", email).eq("campaign", "coaching-invite").order("created_at", desc=True).limit(1).execute().data or []
    if last: out["last_invite"] = last[0]
    sup = _supabase.table("email_suppressions").select("reason,source").eq("email", email).limit(1).execute().data or []
    if sup:
        out["is_suppressed"] = True
        out["suppression_reason"] = sup[0].get("reason")
    return out


@router.post("/clients/{client_id}/provision-portal")
def provision_portal(client_id: int, payload: ProvisionPortalIn, request: Request):
    """Create a users row + own workspace for a coaching client so they can log in
    to mission.tplcollective.ai (admin view) or portal.tplcollective.ai with their own
    creds. The coaching_client.user_id field links them; their workspace stays isolated
    from Joe's data; the coaching API uses /me endpoints scoped by user_id, not workspace."""
    import secrets as _secrets
    from auth import hash_password as _hp

    cc = _supabase.table("coaching_clients").select("*").eq("id", client_id).single().execute().data
    if not cc:
        raise HTTPException(404, "Coaching client not found")
    if cc.get("user_id"):
        return {"already_provisioned": True, "user_id": cc["user_id"]}

    lead = _supabase.table("leads").select("*").eq("id", cc["lead_id"]).single().execute().data or {}
    email = (lead.get("email") or "").strip().lower()
    name = lead.get("name") or ((lead.get("first_name") or "") + " " + (lead.get("last_name") or "")).strip()
    if not email or not is_valid_email(email):
        raise HTTPException(400, "Lead email is missing or invalid")

    # Reuse if a user already exists for this email
    existing = _supabase.table("users").select("id, workspace_id").eq("email", email).execute().data or []
    if existing:
        user_id = existing[0]["id"]
        _supabase.table("coaching_clients").update({"user_id": user_id}).eq("id", client_id).execute()
        return {"reused_existing_user": True, "user_id": user_id}

    # Create the user + workspace
    temp_pwd = _secrets.token_urlsafe(8)
    password_hash = _hp(temp_pwd)
    initials = "".join([p[0] for p in name.split()[:2] if p]).upper() or "AG"

    # Create user first (workspace_id will be set after we make the workspace)
    user_ins = _supabase.table("users").insert({
        "email": email,
        "password_hash": password_hash,
        "name": name,
        "role": "agent",
        "avatar_initials": initials[:2],
    }).execute()
    user_id = user_ins.data[0]["id"]

    # Create the workspace (basic plan — coaching access doesn't need elite/mid)
    ws_ins = _supabase.table("workspaces").insert({
        "owner_user_id": user_id,
        "name": name + "'s Workspace",
        "plan": "basic",
        "settings": {"coaching_only": True},
    }).execute()
    ws_id = ws_ins.data[0]["id"]

    # Update user.workspace_id, link coaching_client.user_id
    _supabase.table("users").update({"workspace_id": ws_id}).eq("id", user_id).execute()
    _supabase.table("coaching_clients").update({"user_id": user_id}).eq("id", client_id).execute()

    # Send invite email with temp password
    if payload.send_email:
        try:
            from main import send_email as _send, load_settings as _load_settings
            settings = _load_settings()
            smtp_cfg = settings.get("smtp", {})
            html = f"""
            <div style='font-family:sans-serif;max-width:560px;margin:0 auto;padding:20px'>
              <h2 style='color:#6c63ff'>Welcome to Coaching with TPL Collective</h2>
              <p>Hey {name.split(' ')[0]} —</p>
              <p>Your coaching portal is ready. Log in here to view your business plan, log daily activity, track your pipeline, and see prep for our calls.</p>
              <p style='background:#f5f5fa;border-radius:8px;padding:14px;font-family:monospace;font-size:13px'>
                <strong>Login:</strong> https://portal.tplcollective.ai<br>
                <strong>Email:</strong> {email}<br>
                <strong>Temporary password:</strong> {temp_pwd}
              </p>
              <p>Change your password after first login. See you on our next call.</p>
              <p>— Joe</p>
            </div>
            """
            _send(smtp_cfg, email, "Your TPL Coaching Portal is Ready", html, from_address="Joe DeSane <joe@tplcollective.co>", campaign="coaching-invite")
        except Exception:
            pass  # Email failure shouldn't block provisioning

    return {"provisioned": True, "user_id": user_id, "workspace_id": ws_id, "temp_password": temp_pwd, "email": email}


# ─── Agent self-service endpoints (/api/coaching/me/*) ───
# Scoped by current_user.sub via coaching_clients.user_id (NOT by workspace).
# Lets an agent in their own isolated workspace still see their coaching record
# which lives in Joe's workspace.

def _my_client(request: Request) -> dict:
    user = getattr(request.state, "user", None) or {}
    user_id = user.get("sub")
    if not user_id:
        raise HTTPException(401, "Not authenticated")
    res = _supabase.table("coaching_clients").select("*").eq("user_id", user_id).limit(1).execute()
    if not res.data:
        raise HTTPException(404, "You are not enrolled as a coaching client")
    return res.data[0]


@router.get("/me")
def my_coaching(request: Request):
    cc = _my_client(request)
    cc = _enrich_client(cc)
    bundle = _ensure_business_plan(cc["id"], datetime.utcnow().year, cc["workspace_id"])
    plan = bundle["plan"]
    em = bundle["economic_model"]
    bm = bundle["budget_model"]
    economic = calc_economic(float(plan.get("gci_target") or 0), em)
    budget = calc_budget(float(plan.get("gci_target") or 0), bm, economic)
    return {
        "client": cc,
        "plan": plan,
        "budget_model": bm,
        "economic_model": em,
        "computed": {"economic": economic, "budget": budget},
    }


@router.get("/me/pipeline")
def my_pipeline(request: Request, entry_type: Optional[str] = None):
    cc = _my_client(request)
    qb = _supabase.table("pipeline_entries").select("*").eq("coaching_client_id", cc["id"])
    if entry_type:
        qb = qb.eq("entry_type", entry_type)
    return qb.order("rating", desc=True).execute().data or []


@router.post("/me/pipeline")
def my_pipeline_create(payload: PipelineEntryIn, request: Request):
    cc = _my_client(request)
    if payload.entry_type not in ("LISTING", "BUYER"):
        raise HTTPException(400, "entry_type must be LISTING or BUYER")
    if payload.rating is not None and not (1 <= payload.rating <= 10):
        raise HTTPException(400, "rating must be 1–10")
    row = {"workspace_id": cc["workspace_id"], "coaching_client_id": cc["id"], **payload.dict(exclude_unset=True)}
    return _supabase.table("pipeline_entries").insert(row).execute().data[0]


@router.patch("/me/pipeline/{entry_id}")
def my_pipeline_update(entry_id: int, payload: PipelineEntryIn, request: Request):
    cc = _my_client(request)
    own = _supabase.table("pipeline_entries").select("id").eq("id", entry_id).eq("coaching_client_id", cc["id"]).execute().data
    if not own:
        raise HTTPException(404, "Not found")
    data = payload.dict(exclude_unset=True)
    res = _supabase.table("pipeline_entries").update(data).eq("id", entry_id).execute()
    return res.data[0]


@router.delete("/me/pipeline/{entry_id}")
def my_pipeline_delete(entry_id: int, request: Request):
    cc = _my_client(request)
    own = _supabase.table("pipeline_entries").select("id").eq("id", entry_id).eq("coaching_client_id", cc["id"]).execute().data
    if not own:
        raise HTTPException(404, "Not found")
    _supabase.table("pipeline_entries").delete().eq("id", entry_id).execute()
    return {"ok": True}


@router.get("/me/activity")
def my_activity(request: Request, days: int = 30):
    cc = _my_client(request)
    return _supabase.table("coaching_activity_logs").select("*").eq("coaching_client_id", cc["id"]).order("log_date", desc=True).limit(max(1, min(days, 365))).execute().data or []


@router.post("/me/activity")
def my_activity_upsert(payload: ActivityLogIn, request: Request):
    cc = _my_client(request)
    log_date = payload.log_date or date.today().isoformat()
    row = {
        "coaching_client_id": cc["id"],
        "log_date": log_date,
        **{k: v for k, v in payload.dict(exclude_unset=True).items() if k != "log_date"},
    }
    return _supabase.table("coaching_activity_logs").upsert(row, on_conflict="coaching_client_id,log_date").execute().data[0]


@router.get("/me/calls")
def my_calls(request: Request):
    cc = _my_client(request)
    return _supabase.table("coaching_calls").select("*").eq("coaching_client_id", cc["id"]).order("scheduled_at", desc=True).limit(50).execute().data or []


@router.get("/me/action-items")
def my_action_items(request: Request, status: Optional[str] = None):
    cc = _my_client(request)
    qb = _supabase.table("coaching_action_items").select("*").eq("coaching_client_id", cc["id"]).eq("owner", "AGENT")
    if status:
        qb = qb.eq("status", status)
    return qb.order("due_date", desc=False).execute().data or []


@router.patch("/me/action-items/{item_id}")
def my_action_item_status(item_id: int, payload: ActionItemUpdate, request: Request):
    """Agents can mark items COMPLETED/MISSED/OPEN, but can't edit text/measurement/due/owner/tag."""
    cc = _my_client(request)
    own = _supabase.table("coaching_action_items").select("id").eq("id", item_id).eq("coaching_client_id", cc["id"]).eq("owner", "AGENT").execute().data
    if not own:
        raise HTTPException(404, "Not found or not yours")
    if not payload.status:
        raise HTTPException(400, "status is required")
    data = {"status": payload.status}
    if payload.status == "COMPLETED":
        data["completed_at"] = datetime.utcnow().isoformat()
    return _supabase.table("coaching_action_items").update(data).eq("id", item_id).execute().data[0]


@router.get("/me/brief")
def my_brief(request: Request):
    """The agent sees the same pre-call brief their coach uses."""
    cc = _my_client(request)
    last = _supabase.table("coaching_calls").select("id").eq("coaching_client_id", cc["id"]).order("scheduled_at", desc=True).limit(1).execute()
    if last.data:
        return _build_brief(cc["id"], last.data[0]["id"], cc["workspace_id"])
    return _build_brief_no_call(cc["id"], cc["workspace_id"])


class OnboardIn(BaseModel):
    """CTE-style detailed goal-setting form payload (LPT-only).
    All sections optional so the wizard can save partial progress.

    Drives from net_income_goal (Step 2) and derives GCI via LPT comp-plan math.
    LPT comp plan in client.lpt_comp_plan auto-populates cap / TC fee / split %
    so the agent never types broker-specific numbers we already know.
    """
    # Identity / Vision (Step 1)
    client: Optional[dict] = None     # CoachingClientUpdate-shape, plus: big_why, team_or_individual, team_name
    # Income Goal (Step 2) — net_income_goal is the primary input; GCI is derived
    plan: Optional[dict] = None       # {net_income_goal, gci_target (computed), notes}
    pct_listing_income: Optional[float] = None   # 0-100, % income from listings
    # Deal Economics (Step 3)
    economic: Optional[dict] = None   # {seller_avg_sale_price, buyer_avg_sale_price, commission_rate_listing, commission_rate_buyer, listings_close_pct, buyers_close_pct, listing_appt_to_list_pct, buyer_appt_to_work_pct}
    # Money Plan (Step 4) — LPT-only, no royalty / non-LPT fields
    budget: Optional[dict] = None     # {income_tax_pct, charity_pct, retirement_pct, avg_net_commission_per_close (computed)}
    # Lead-Gen Methods (Step 5) — checklist of how the agent prospects
    lead_gen_methods: Optional[list] = None
    # Recruiting / Rev Share (Step 6, LPT only)
    recruiting: Optional[dict] = None


@router.post("/me/onboard")
def my_onboard(payload: OnboardIn, request: Request):
    """Agent self-service: fill in the LPT-aware goal-setting form.
    Drives from net_income_goal; derives GCI / closings / cost-of-sale via LPT comp-plan math.
    Saves partial progress safely — every section is optional."""
    cc = _my_client(request)

    # ── Step 1: Identity / Vision (LPT-only) ──
    client_data = dict(payload.client or {})
    # Force brokerage = LPT for coaching clients (this platform is LPT-specific)
    if client_data:
        client_data.setdefault("brokerage", "LPT")
        if "avg_commission_rate" in client_data and client_data["avg_commission_rate"] is not None:
            client_data["avg_commission_rate"] = _normalize_pct(client_data["avg_commission_rate"])
        _supabase.table("coaching_clients").update(client_data).eq("id", cc["id"]).execute()

    yr = datetime.utcnow().year
    bundle = _ensure_business_plan(cc["id"], yr, cc["workspace_id"])
    plan_id = bundle["plan"]["id"]

    # ── Step 3: Deal Economics (saved before Step 2 so we have ASP for derivation) ──
    em_update = dict(payload.economic or {})
    if payload.pct_listing_income is not None:
        em_update["seller_pct"] = _normalize_pct(payload.pct_listing_income)
    for k in ("seller_pct", "commission_rate", "commission_rate_listing", "commission_rate_buyer",
              "listings_close_pct", "buyers_close_pct",
              "listing_appt_to_list_pct", "buyer_appt_to_work_pct"):
        if k in em_update and em_update[k] is not None:
            em_update[k] = _normalize_pct(em_update[k])
    if em_update.get("commission_rate_listing") and not em_update.get("commission_rate"):
        em_update["commission_rate"] = em_update["commission_rate_listing"]
    if em_update:
        _supabase.table("economic_models").update(em_update).eq("business_plan_id", plan_id).execute()

    # ── Step 2: Income Goal — derive GCI from net_income_goal + comp plan + avg deal ──
    plan_data = dict(payload.plan or {})
    # Normalize: net_income_goal is the input, gci_target is derived
    net_goal = plan_data.get("net_income_goal")
    if net_goal is not None and float(net_goal) > 0:
        # Pull current avg deal economics + comp plan (whatever's freshest)
        em = _supabase.table("economic_models").select("*").eq("business_plan_id", plan_id).single().execute().data or {}
        cc_now = _supabase.table("coaching_clients").select("lpt_comp_plan").eq("id", cc["id"]).single().execute().data or {}
        comp_plan = cc_now.get("lpt_comp_plan") or "BROKERAGE_PARTNER"
        # Weighted avg deal GCI = listing_pct × (listing_asp × listing_comm) + buyer_pct × (buyer_asp × buyer_comm)
        listing_pct = float(em.get("seller_pct") or 0.5)
        listing_asp = float(em.get("seller_avg_sale_price") or 0)
        buyer_asp = float(em.get("buyer_avg_sale_price") or 0)
        listing_comm = float(em.get("commission_rate_listing") or em.get("commission_rate") or 0.025)
        buyer_comm = float(em.get("commission_rate_buyer") or em.get("commission_rate") or 0.025)
        avg_gci_per_deal = (listing_pct * listing_asp * listing_comm) + ((1 - listing_pct) * buyer_asp * buyer_comm)
        if avg_gci_per_deal > 0:
            derived = compute_gci_from_net_income(float(net_goal), avg_gci_per_deal, comp_plan)
            if derived:
                plan_data["gci_target"] = derived["gci_target"]
                # Also update budget_models.avg_net_commission_per_close so downstream math has it
                _supabase.table("budget_models").update({
                    "avg_net_commission_per_close": derived["avg_net_per_close"],
                }).eq("business_plan_id", plan_id).execute()
    if plan_data:
        _supabase.table("business_plans").update(plan_data).eq("id", plan_id).execute()

    # ── Step 4: Money Plan (LPT auto-populates cap + per-txn fee + TC) ──
    bm_update = dict(payload.budget or {})
    for k in ("income_tax_pct", "charity_pct", "retirement_pct"):
        if k in bm_update and bm_update[k] is not None:
            bm_update[k] = _normalize_pct(bm_update[k])
    # LPT-only: auto-set cap, zero out royalty, never let agent type these
    cc_for_plan = _supabase.table("coaching_clients").select("lpt_comp_plan").eq("id", cc["id"]).single().execute().data or {}
    plan_details = lpt_plan_details(cc_for_plan.get("lpt_comp_plan"))
    if plan_details:
        bm_update["paid_to_brokerage"] = plan_details["cap"]
        bm_update["royalty_pct"] = 0
        bm_update["royalty_cap"] = 0
    if bm_update:
        _supabase.table("budget_models").update(bm_update).eq("business_plan_id", plan_id).execute()

    # ── Step 5: Lead-Gen Methods (replaces old numeric activity goals) ──
    if payload.lead_gen_methods is not None:
        _supabase.table("business_plans").update({
            "lead_gen_methods": payload.lead_gen_methods,
        }).eq("id", plan_id).execute()

    # ── Step 6: Recruiting / Rev Share ──
    rp = dict(payload.recruiting or {})
    for k in ("pct_brokerage_partner", "cap_hit_rate"):
        if k in rp and rp[k] is not None:
            rp[k] = _normalize_pct(rp[k])
    if rp:
        rp["business_plan_id"] = plan_id
        _supabase.table("recruiting_plans").upsert(rp, on_conflict="business_plan_id").execute()

    return {"ok": True}


@router.get("/me/derive-gci")
def my_derive_gci(request: Request, net_income_goal: float, pct_listing_income: float = 50):
    """Live preview helper for Step 2: returns the GCI that would be needed to hit
    the entered net-income goal, given the agent's current LPT comp plan + avg deal."""
    cc = _my_client(request)
    yr = datetime.utcnow().year
    bundle = _ensure_business_plan(cc["id"], yr, cc["workspace_id"])
    plan_id = bundle["plan"]["id"]
    em = _supabase.table("economic_models").select("*").eq("business_plan_id", plan_id).single().execute().data or {}
    cc_now = _supabase.table("coaching_clients").select("lpt_comp_plan").eq("id", cc["id"]).single().execute().data or {}
    comp_plan = cc_now.get("lpt_comp_plan") or "BROKERAGE_PARTNER"
    listing_pct = pct_listing_income / 100.0
    listing_asp = float(em.get("seller_avg_sale_price") or 0)
    buyer_asp = float(em.get("buyer_avg_sale_price") or 0)
    listing_comm = float(em.get("commission_rate_listing") or em.get("commission_rate") or 0.025)
    buyer_comm = float(em.get("commission_rate_buyer") or em.get("commission_rate") or 0.025)
    avg_gci = (listing_pct * listing_asp * listing_comm) + ((1 - listing_pct) * buyer_asp * buyer_comm)
    if avg_gci <= 0:
        return {"ok": False, "reason": "Need avg sale price + commission % from Step 3 first"}
    return compute_gci_from_net_income(float(net_income_goal), avg_gci, comp_plan)


@router.get("/lpt-plan-details")
def lpt_plan_details_endpoint(comp_plan: str, request: Request):
    """Returns the LPT comp plan constants so the frontend can show / preview them."""
    d = lpt_plan_details(comp_plan)
    if not d:
        raise HTTPException(400, "Unknown comp plan — must be BROKERAGE_PARTNER or BUSINESS_BUILDER")
    return d


@router.get("/me/activity-goals")
def my_activity_goals(request: Request):
    """Return the agent's per-category activity goals (used for pace calc against actuals)."""
    cc = _my_client(request)
    yr = datetime.utcnow().year
    bundle = _ensure_business_plan(cc["id"], yr, cc["workspace_id"])
    res = _supabase.table("activity_goals").select("*").eq("business_plan_id", bundle["plan"]["id"]).execute()
    return res.data[0] if res.data else {}


# ─── Scorecard rollups (CTE Scorecards tab equivalent) ───

def _scorecard_for_period(client_id: int, start_date: date, end_date: date) -> dict:
    """Sum the activity log fields between two dates (inclusive). Returns the rollup
    plus computed conversion percentages."""
    res = _supabase.table("coaching_activity_logs").select("*") \
        .eq("coaching_client_id", client_id) \
        .gte("log_date", start_date.isoformat()) \
        .lte("log_date", end_date.isoformat()) \
        .execute()
    rows = res.data or []
    days_logged = len(rows)
    fields = [
        "dials", "contacts_made", "nurtures",
        "listing_appts_set", "listing_appts_held", "listings_signed",
        "buyer_appts_set", "buyer_appts_held", "buyers_signed",
        "showings", "open_houses_held",
        "appts_set", "appts_held", "hours_prospected",
    ]
    totals = {f: 0 for f in fields}
    for r in rows:
        for f in fields:
            v = r.get(f) or 0
            totals[f] += float(v) if f == "hours_prospected" else int(v)
    # Derived conversions
    contact_to_appt = 0
    if totals["contacts_made"] > 0:
        all_appts_set = totals["listing_appts_set"] + totals["buyer_appts_set"] + totals["appts_set"]
        contact_to_appt = round((all_appts_set / totals["contacts_made"]) * 100, 1)
    set_to_held = 0
    all_set = totals["listing_appts_set"] + totals["buyer_appts_set"] + totals["appts_set"]
    all_held = totals["listing_appts_held"] + totals["buyer_appts_held"] + totals["appts_held"]
    if all_set > 0:
        set_to_held = round((all_held / all_set) * 100, 1)
    contacts_per_hour = 0
    if totals["hours_prospected"] > 0:
        contacts_per_hour = round(totals["contacts_made"] / totals["hours_prospected"], 1)
    return {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "days_logged": days_logged,
        "totals": totals,
        "contact_to_appt_pct": contact_to_appt,
        "set_to_held_pct": set_to_held,
        "contacts_per_hour": contacts_per_hour,
    }


def _scorecard_full(client_id: int, year: int, workspace_id: int) -> dict:
    """CTE-style scorecard: this week, this month, last 30 days, YTD, plus monthly grid Jan-Dec."""
    today = date.today()
    from datetime import timedelta as _td
    # This week (Mon-Sun)
    week_start = today - _td(days=today.weekday())
    week_end = week_start + _td(days=6)
    # This month
    month_start = today.replace(day=1)
    if today.month == 12:
        month_end = today.replace(year=today.year+1, month=1, day=1) - _td(days=1)
    else:
        month_end = today.replace(month=today.month+1, day=1) - _td(days=1)
    # Last 30 days
    last30_start = today - _td(days=29)
    # YTD
    ytd_start = date(year, 1, 1)
    ytd_end = date(year, 12, 31)

    bundle = _ensure_business_plan(client_id, year, workspace_id)
    plan_id = bundle["plan"]["id"]
    goals_resp = _supabase.table("activity_goals").select("*").eq("business_plan_id", plan_id).execute()
    goals = goals_resp.data[0] if goals_resp.data else {}

    # Monthly grid
    monthly = []
    for m in range(1, 13):
        ms = date(year, m, 1)
        if m == 12:
            me = date(year, 12, 31)
        else:
            me = date(year, m+1, 1) - _td(days=1)
        monthly.append({"month": m, **_scorecard_for_period(client_id, ms, me)})

    return {
        "year": year,
        "this_week": _scorecard_for_period(client_id, week_start, week_end),
        "this_month": _scorecard_for_period(client_id, month_start, month_end),
        "last_30_days": _scorecard_for_period(client_id, last30_start, today),
        "ytd": _scorecard_for_period(client_id, ytd_start, ytd_end),
        "monthly": monthly,
        "goals": goals,
    }


@router.get("/me/scorecard")
def my_scorecard(request: Request, year: Optional[int] = None):
    cc = _my_client(request)
    yr = year or datetime.utcnow().year
    return _scorecard_full(cc["id"], yr, cc["workspace_id"])


# ─── CEO Summary + Whiteboard (transactions by status) ───

def _ceo_summary(client_id: int, year: int) -> dict:
    """CTE CEO Summary equivalent: pipeline counts by status + closed YTD totals + monthly grid."""
    today = date.today()
    rows = _supabase.table("pipeline_entries").select("*").eq("coaching_client_id", client_id).execute().data or []

    # Counts by status
    by_status = {"PRE_SIGNED":0, "ACTIVE":0, "PENDING":0, "CLOSED":0, "EXPIRED":0, "WITHDRAWN":0, "CANCELLED":0}
    listing_active = 0; buyer_active = 0
    listing_pending = 0; buyer_pending = 0
    closed_ytd = []
    pending_volume = 0
    pending_gci = 0
    active_listings_volume = 0
    active_listings_gci = 0  # projected GCI on active listings

    for r in rows:
        st = (r.get("status") or "PRE_SIGNED").upper()
        by_status[st] = by_status.get(st, 0) + 1
        et = r.get("entry_type")
        if st == "ACTIVE":
            if et == "LISTING":
                listing_active += 1
                active_listings_volume += float(r.get("closing_price") or 0)  # price = list price for active
                active_listings_gci += float(r.get("gross_commission") or 0)
            elif et == "BUYER":
                buyer_active += 1
        elif st == "PENDING":
            if et == "LISTING": listing_pending += 1
            else: buyer_pending += 1
            pending_volume += float(r.get("closing_price") or 0)
            pending_gci += float(r.get("gross_commission") or 0)
        elif st == "CLOSED":
            cd = r.get("closed_date") or ""
            if cd.startswith(str(year)):
                closed_ytd.append(r)

    # Closed YTD totals
    closed_units = len(closed_ytd)
    closed_volume = sum(float(r.get("closing_price") or 0) for r in closed_ytd)
    closed_gci = sum(float(r.get("gross_commission") or 0) for r in closed_ytd)
    closed_net = sum(float(r.get("net_commission") or 0) for r in closed_ytd)

    # Monthly closed grid (Jan-Dec)
    monthly_closed = []
    for m in range(1, 13):
        units = sum(1 for r in closed_ytd if (r.get("closed_date") or "").startswith(f"{year}-{m:02d}"))
        vol = sum(float(r.get("closing_price") or 0) for r in closed_ytd if (r.get("closed_date") or "").startswith(f"{year}-{m:02d}"))
        gci = sum(float(r.get("gross_commission") or 0) for r in closed_ytd if (r.get("closed_date") or "").startswith(f"{year}-{m:02d}"))
        monthly_closed.append({"month": m, "units": units, "volume": round(vol,2), "gci": round(gci,2)})

    return {
        "year": year,
        "as_of": today.isoformat(),
        "by_status": by_status,
        "listing_active": listing_active,
        "buyer_active": buyer_active,
        "listing_pending": listing_pending,
        "buyer_pending": buyer_pending,
        "active_listings_volume": round(active_listings_volume, 2),
        "active_listings_gci": round(active_listings_gci, 2),
        "pending_volume": round(pending_volume, 2),
        "pending_gci": round(pending_gci, 2),
        "closed_ytd_units": closed_units,
        "closed_ytd_volume": round(closed_volume, 2),
        "closed_ytd_gci": round(closed_gci, 2),
        "closed_ytd_net_gci": round(closed_net, 2),
        "monthly_closed": monthly_closed,
    }


@router.get("/me/ceo-summary")
def my_ceo_summary(request: Request, year: Optional[int] = None):
    cc = _my_client(request)
    return _ceo_summary(cc["id"], year or datetime.utcnow().year)


@router.get("/clients/{client_id}/ceo-summary")
def coach_ceo_summary(client_id: int, request: Request, year: Optional[int] = None):
    return _ceo_summary(client_id, year or datetime.utcnow().year)


def _whiteboard(client_id: int) -> dict:
    """CTE Whiteboard equivalent: active listings with DOM, expiring soon, expiring next 30 days."""
    today = date.today()
    rows = _supabase.table("pipeline_entries").select("*").eq("coaching_client_id", client_id).eq("entry_type", "LISTING").execute().data or []
    active = []
    over_90 = []
    expiring_30 = []
    for r in rows:
        st = (r.get("status") or "PRE_SIGNED").upper()
        if st != "ACTIVE":
            continue
        ld = r.get("list_date")
        ed = r.get("expiration_date")
        dom = None
        if ld:
            try:
                d = datetime.strptime(ld, "%Y-%m-%d").date()
                dom = (today - d).days
            except Exception:
                pass
        days_to_expiry = None
        if ed:
            try:
                d = datetime.strptime(ed, "%Y-%m-%d").date()
                days_to_expiry = (d - today).days
            except Exception:
                pass
        item = {**r, "dom": dom, "days_to_expiry": days_to_expiry}
        active.append(item)
        if dom is not None and dom > 90:
            over_90.append(item)
        if days_to_expiry is not None and 0 <= days_to_expiry <= 30:
            expiring_30.append(item)
    # Sort active by DOM desc (oldest at top — most attention needed)
    active.sort(key=lambda x: (x.get("dom") or 0), reverse=True)
    expiring_30.sort(key=lambda x: x.get("days_to_expiry") or 0)
    return {
        "as_of": today.isoformat(),
        "active": active,
        "over_90_days": over_90,
        "expiring_in_30_days": expiring_30,
    }


@router.get("/me/whiteboard")
def my_whiteboard(request: Request):
    cc = _my_client(request)
    return _whiteboard(cc["id"])


@router.get("/clients/{client_id}/whiteboard")
def coach_whiteboard(client_id: int, request: Request):
    return _whiteboard(client_id)


# ════════════════════════════════════════════════════════════
# Reviews (quarterly / semi-annual / annual checkpoints)
# ════════════════════════════════════════════════════════════
# A "review" snapshots the agent's full state at a point in time:
# plan + computed numbers + scorecard + CEO summary + recent activity.
# Plus the coach's narrative (reflections + focus areas for next period).

class ReviewIn(BaseModel):
    review_type: str            # QUARTERLY | SEMI_ANNUAL | ANNUAL
    review_date: Optional[str] = None
    reflections: Optional[str] = None
    focus_areas_next: Optional[str] = None


class ReviewUpdate(BaseModel):
    review_date: Optional[str] = None
    reflections: Optional[str] = None
    focus_areas_next: Optional[str] = None
    recapture: Optional[bool] = False   # if true, re-snapshot captured_data from current state


def _capture_review_data(client_id: int, workspace_id: int, year: int) -> dict:
    """Build the JSONB snapshot — same data the coach sees on a brief, plus YTD scorecard + CEO summary."""
    cc = _supabase.table("coaching_clients").select("*").eq("id", client_id).single().execute().data or {}
    bundle = _ensure_business_plan(client_id, year, workspace_id)
    plan = bundle["plan"]
    em = bundle["economic_model"]
    bm = bundle["budget_model"]
    gci = float(plan.get("gci_target") or 0)
    economic = calc_economic(gci, em)
    budget = calc_budget(gci, bm, economic)
    scorecard = _scorecard_full(client_id, year, workspace_id)
    ceo = _ceo_summary(client_id, year)
    return {
        "client_meta": {k: cc.get(k) for k in ["brokerage", "lpt_comp_plan", "call_cadence", "market_city", "market_state", "big_why"]},
        "plan": {k: plan.get(k) for k in ["year", "gci_target", "notes"]},
        "computed_economic": economic,
        "computed_budget": budget,
        "scorecard": scorecard,
        "ceo_summary": ceo,
    }


@router.get("/clients/{client_id}/reviews")
def list_reviews(client_id: int, request: Request, review_type: Optional[str] = None):
    qb = _db("review_snapshots").select("*").eq("coaching_client_id", client_id)
    if review_type:
        qb = qb.eq("review_type", review_type)
    return qb.order("review_date", desc=True).execute().data or []


@router.get("/reviews/{review_id}")
def get_review(review_id: int, request: Request):
    res = _db("review_snapshots").select("*").eq("id", review_id).execute()
    if not res.data:
        raise HTTPException(404, "Review not found")
    return res.data[0]


@router.post("/clients/{client_id}/reviews")
def create_review(client_id: int, payload: ReviewIn, request: Request):
    """Capture a review snapshot at a checkpoint."""
    workspace_id = _ws(request)
    if payload.review_type not in ("QUARTERLY", "SEMI_ANNUAL", "ANNUAL"):
        raise HTTPException(400, "review_type must be QUARTERLY / SEMI_ANNUAL / ANNUAL")
    review_date = payload.review_date or date.today().isoformat()
    yr = int(review_date[:4]) if review_date else datetime.utcnow().year
    captured = _capture_review_data(client_id, workspace_id, yr)
    row = {
        "workspace_id": workspace_id,
        "coaching_client_id": client_id,
        "review_type": payload.review_type,
        "review_date": review_date,
        "captured_data": captured,
        "reflections": payload.reflections,
        "focus_areas_next": payload.focus_areas_next,
    }
    return _supabase.table("review_snapshots").insert(row).execute().data[0]


@router.patch("/reviews/{review_id}")
def update_review(review_id: int, payload: ReviewUpdate, request: Request):
    workspace_id = _ws(request)
    cur = _supabase.table("review_snapshots").select("*").eq("id", review_id).single().execute().data
    if not cur:
        raise HTTPException(404, "Review not found")
    data = payload.dict(exclude_unset=True)
    if data.pop("recapture", False):
        yr = int((cur.get("review_date") or "")[:4] or datetime.utcnow().year)
        data["captured_data"] = _capture_review_data(cur["coaching_client_id"], workspace_id, yr)
    if not data:
        raise HTTPException(400, "No fields to update")
    return _supabase.table("review_snapshots").update(data).eq("id", review_id).execute().data[0]


@router.delete("/reviews/{review_id}")
def delete_review(review_id: int, request: Request):
    _supabase.table("review_snapshots").delete().eq("id", review_id).execute()
    return {"ok": True}


# ════════════════════════════════════════════════════════════
# Database Touch Tracker (contact_touches)
# ════════════════════════════════════════════════════════════
# MREA model: agents must touch each Met contact 12x/year and each
# Haven't-Met contact differently. We track every touch with date +
# method, then compute who's overdue.
#
# Schema (already in place from session 1):
#   contact_touches: person_name, person_email, person_phone, contact_type
#                    (MET | HAVENT_MET), touch_date, touch_method, notes
#
# We track the *touches*, not the contacts as separate rows. The
# unique-person dimension comes from grouping by (person_name + email).

class TouchIn(BaseModel):
    person_name: str
    person_email: Optional[str] = None
    person_phone: Optional[str] = None
    contact_type: Optional[str] = "MET"   # MET | HAVENT_MET
    touch_date: Optional[str] = None
    touch_method: Optional[str] = None    # CALL | TEXT | EMAIL | EVENT | MAIL | SOCIAL | IN_PERSON | OTHER
    notes: Optional[str] = None


class TouchUpdate(BaseModel):
    person_name: Optional[str] = None
    person_email: Optional[str] = None
    person_phone: Optional[str] = None
    contact_type: Optional[str] = None
    touch_date: Optional[str] = None
    touch_method: Optional[str] = None
    notes: Optional[str] = None


def _touches_database(client_id: int) -> dict:
    """Group touches by person + compute overdue status.
    A "person" is unique on (name lower + email lower) since email may be missing.
    Met target = 12 touches/year (1/month). Haven't-Met = MREA doesn't fix a cadence,
    so we use a simpler heuristic: overdue if no touch in last 90 days."""
    today = date.today()
    rows = _supabase.table("contact_touches").select("*").eq("coaching_client_id", client_id).order("touch_date", desc=True).execute().data or []
    by_person = {}
    for r in rows:
        key = ((r.get("person_name") or "").strip().lower(), (r.get("person_email") or "").strip().lower())
        if not key[0] and not key[1]: continue
        if key not in by_person:
            by_person[key] = {
                "person_name": r.get("person_name"),
                "person_email": r.get("person_email"),
                "person_phone": r.get("person_phone"),
                "contact_type": r.get("contact_type") or "MET",
                "touches": [],
                "last_touch": None,
                "ytd_count": 0,
                "days_since": None,
                "overdue": False,
            }
        person = by_person[key]
        person["touches"].append(r)
        # Update aggregates
        td = r.get("touch_date")
        if td:
            try:
                d = datetime.strptime(td, "%Y-%m-%d").date()
                if person["last_touch"] is None or d > datetime.strptime(person["last_touch"], "%Y-%m-%d").date():
                    person["last_touch"] = td
                if d.year == today.year:
                    person["ytd_count"] += 1
            except Exception:
                pass
    # Compute days_since + overdue for each person
    overdue_count = 0
    for key, p in by_person.items():
        if p["last_touch"]:
            try:
                d = datetime.strptime(p["last_touch"], "%Y-%m-%d").date()
                p["days_since"] = (today - d).days
            except Exception:
                pass
        # Overdue rules:
        # MET: <12 YTD AND days_since > 30 (off pace for 12/yr)
        # HAVENT_MET: days_since > 90
        if p["contact_type"] == "MET":
            if p["days_since"] is None or p["days_since"] > 30:
                p["overdue"] = True
        else:
            if p["days_since"] is None or p["days_since"] > 90:
                p["overdue"] = True
        if p["overdue"]: overdue_count += 1

    persons = list(by_person.values())
    # Sort: overdue first (most days_since at top), then by ytd_count desc
    persons.sort(key=lambda p: (not p["overdue"], -(p["days_since"] or 0)))

    # Summary stats
    met_count = sum(1 for p in persons if p["contact_type"] == "MET")
    havent_count = sum(1 for p in persons if p["contact_type"] == "HAVENT_MET")
    return {
        "as_of": today.isoformat(),
        "total_persons": len(persons),
        "met_count": met_count,
        "havent_met_count": havent_count,
        "overdue_count": overdue_count,
        "persons": persons,
    }


@router.get("/clients/{client_id}/touches")
def list_touches(client_id: int, request: Request, limit: int = 200):
    """Raw touch history (most recent first)."""
    res = _db("contact_touches").select("*").eq("coaching_client_id", client_id).order("touch_date", desc=True).limit(limit).execute()
    return res.data or []


@router.get("/clients/{client_id}/database")
def coach_database(client_id: int, request: Request):
    """Grouped-by-person view with overdue computation."""
    return _touches_database(client_id)


@router.post("/clients/{client_id}/touches")
def create_touch(client_id: int, payload: TouchIn, request: Request):
    row = {
        "coaching_client_id": client_id,
        "person_name": payload.person_name,
        "person_email": payload.person_email,
        "person_phone": payload.person_phone,
        "contact_type": payload.contact_type or "MET",
        "touch_date": payload.touch_date or date.today().isoformat(),
        "touch_method": payload.touch_method,
        "notes": payload.notes,
    }
    return _supabase.table("contact_touches").insert(row).execute().data[0]


@router.patch("/touches/{touch_id}")
def update_touch(touch_id: int, payload: TouchUpdate, request: Request):
    data = payload.dict(exclude_unset=True)
    if not data:
        raise HTTPException(400, "No fields to update")
    res = _supabase.table("contact_touches").update(data).eq("id", touch_id).execute()
    if not res.data:
        raise HTTPException(404, "Touch not found")
    return res.data[0]


@router.delete("/touches/{touch_id}")
def delete_touch(touch_id: int, request: Request):
    _supabase.table("contact_touches").delete().eq("id", touch_id).execute()
    return {"ok": True}


# Agent self-service
@router.get("/me/touches")
def my_touches(request: Request, limit: int = 200):
    cc = _my_client(request)
    res = _supabase.table("contact_touches").select("*").eq("coaching_client_id", cc["id"]).order("touch_date", desc=True).limit(limit).execute()
    return res.data or []


@router.get("/me/database")
def my_database(request: Request):
    cc = _my_client(request)
    return _touches_database(cc["id"])


@router.post("/me/touches")
def my_create_touch(payload: TouchIn, request: Request):
    cc = _my_client(request)
    row = {
        "coaching_client_id": cc["id"],
        "person_name": payload.person_name,
        "person_email": payload.person_email,
        "person_phone": payload.person_phone,
        "contact_type": payload.contact_type or "MET",
        "touch_date": payload.touch_date or date.today().isoformat(),
        "touch_method": payload.touch_method,
        "notes": payload.notes,
    }
    return _supabase.table("contact_touches").insert(row).execute().data[0]


@router.patch("/me/touches/{touch_id}")
def my_update_touch(touch_id: int, payload: TouchUpdate, request: Request):
    cc = _my_client(request)
    own = _supabase.table("contact_touches").select("id").eq("id", touch_id).eq("coaching_client_id", cc["id"]).execute().data
    if not own:
        raise HTTPException(404, "Not found or not yours")
    data = payload.dict(exclude_unset=True)
    if not data:
        raise HTTPException(400, "No fields to update")
    return _supabase.table("contact_touches").update(data).eq("id", touch_id).execute().data[0]


@router.delete("/me/touches/{touch_id}")
def my_delete_touch(touch_id: int, request: Request):
    cc = _my_client(request)
    own = _supabase.table("contact_touches").select("id").eq("id", touch_id).eq("coaching_client_id", cc["id"]).execute().data
    if not own:
        raise HTTPException(404, "Not found or not yours")
    _supabase.table("contact_touches").delete().eq("id", touch_id).execute()
    return {"ok": True}


# ════════════════════════════════════════════════════════════
# Perfect Week scheduler
# ════════════════════════════════════════════════════════════
# perfect_weeks.schedule is JSONB: { "MON": { "BEFORE_8": "...", "MORNING": "...", "AFTERNOON": "...", "EVENING": "..." }, ... }

PERFECT_WEEK_DEFAULT_TEMPLATE = {
    "MON": {"BEFORE_8": "Mindset / workout", "MORNING": "Lead Gen — 50 dials", "AFTERNOON": "Listing appts + Showings", "EVENING": "Buyer consults / family"},
    "TUE": {"BEFORE_8": "Mindset / workout", "MORNING": "Lead Gen — 50 dials", "AFTERNOON": "Listing appts + Showings", "EVENING": "Buyer consults / family"},
    "WED": {"BEFORE_8": "Mindset / workout", "MORNING": "Lead Gen — 50 dials", "AFTERNOON": "Listing appts + Showings", "EVENING": "Mid-week review / coaching call"},
    "THU": {"BEFORE_8": "Mindset / workout", "MORNING": "Lead Gen — 50 dials", "AFTERNOON": "Listing appts + Showings", "EVENING": "Buyer consults / family"},
    "FRI": {"BEFORE_8": "Mindset / workout", "MORNING": "Lead Gen — 50 dials", "AFTERNOON": "Listing appts + Showings", "EVENING": "Closings / week wrap"},
    "SAT": {"BEFORE_8": "", "MORNING": "Open House (if listing active)", "AFTERNOON": "Buyer showings", "EVENING": "Personal"},
    "SUN": {"BEFORE_8": "", "MORNING": "Family / faith", "AFTERNOON": "Plan the week (4-1-1)", "EVENING": "Set tomorrow's intention"},
}


class PerfectWeekIn(BaseModel):
    schedule: Optional[dict] = None
    template_name: Optional[str] = None


def _ensure_perfect_week(client_id: int) -> dict:
    res = _supabase.table("perfect_weeks").select("*").eq("coaching_client_id", client_id).execute()
    if res.data:
        return res.data[0]
    ins = _supabase.table("perfect_weeks").insert({
        "coaching_client_id": client_id,
        "schedule": PERFECT_WEEK_DEFAULT_TEMPLATE,
        "template_name": "Default 50-dial week",
    }).execute()
    return ins.data[0]


@router.get("/clients/{client_id}/perfect-week")
def get_perfect_week(client_id: int, request: Request):
    return _ensure_perfect_week(client_id)


@router.put("/clients/{client_id}/perfect-week")
def update_perfect_week(client_id: int, payload: PerfectWeekIn, request: Request):
    pw = _ensure_perfect_week(client_id)
    data = payload.dict(exclude_unset=True)
    if not data:
        raise HTTPException(400, "No fields to update")
    return _supabase.table("perfect_weeks").update(data).eq("id", pw["id"]).execute().data[0]


@router.get("/me/perfect-week")
def my_perfect_week(request: Request):
    cc = _my_client(request)
    return _ensure_perfect_week(cc["id"])


@router.put("/me/perfect-week")
def my_update_perfect_week(payload: PerfectWeekIn, request: Request):
    cc = _my_client(request)
    pw = _ensure_perfect_week(cc["id"])
    data = payload.dict(exclude_unset=True)
    if not data:
        raise HTTPException(400, "No fields to update")
    return _supabase.table("perfect_weeks").update(data).eq("id", pw["id"]).execute().data[0]


# ════════════════════════════════════════════════════════════
# Monthly Financial Statement (CTE Financial Statement tab)
# ════════════════════════════════════════════════════════════
# 12 rows per business plan. JSONB lines for flexibility. Actuals
# auto-populate income from closed pipeline_entries (Listing Income +
# Buyer Income); coach/agent enters cost_of_sales + operating_expenses
# manually. Computed totals returned alongside.

DEFAULT_INCOME_LINES = {
    "Listing Income": 0,
    "Buyer Income": 0,
    "Referral Income": 0,
    "Lease Income": 0,
    "Other Income": 0,
    "Transaction Fee": 0,
    "Bonus": 0,
}

DEFAULT_COGS_LINES = {
    "Split to Office": 0,
    "Royalty Paid": 0,
    "Referral Fees Paid": 0,
    "Listing Agents": 0,
    "Buyer Agents": 0,
    "E&O": 0,
    "ISA": 0,
    "Showing Assistants": 0,
    "Admin Fee": 0,
}

DEFAULT_OPEX_LINES = {
    "Marketing / Advertising": 0,
    "Technology / Software": 0,
    "Phone": 0,
    "Auto / Gas": 0,
    "Education / Coaching / Travel": 0,
    "Office / Rent": 0,
    "Supplies": 0,
}


class FinancialMonthIn(BaseModel):
    income: Optional[dict] = None
    cost_of_sales: Optional[dict] = None
    operating_expenses: Optional[dict] = None
    notes: Optional[str] = None


def _ensure_financial_months(plan_id: int) -> list:
    """Returns the 12 monthly_financials rows, auto-creating any missing months with defaults."""
    rows = _supabase.table("monthly_financials").select("*").eq("business_plan_id", plan_id).order("month").execute().data or []
    by_month = {r["month"]: r for r in rows}
    out = []
    for m in range(1, 13):
        if m in by_month:
            out.append(by_month[m])
        else:
            ins = _supabase.table("monthly_financials").insert({
                "business_plan_id": plan_id, "month": m,
                "income": DEFAULT_INCOME_LINES.copy(),
                "cost_of_sales": DEFAULT_COGS_LINES.copy(),
                "operating_expenses": DEFAULT_OPEX_LINES.copy(),
            }).execute()
            out.append(ins.data[0])
    return out


def _financials_with_actuals(client_id: int, year: int, workspace_id: int) -> dict:
    """Return the 12-month grid with auto-populated actuals from closed transactions and computed totals."""
    bundle = _ensure_business_plan(client_id, year, workspace_id)
    plan_id = bundle["plan"]["id"]
    months = _ensure_financial_months(plan_id)

    # Pull all closed pipeline entries for the year, group by month + entry_type
    closed = _supabase.table("pipeline_entries").select("entry_type,closed_date,gross_commission").eq("coaching_client_id", client_id).eq("closed", True).execute().data or []
    listing_actual = {m: 0 for m in range(1, 13)}
    buyer_actual = {m: 0 for m in range(1, 13)}
    for r in closed:
        cd = r.get("closed_date") or ""
        if not cd.startswith(str(year)): continue
        try:
            m = int(cd.split("-")[1])
        except Exception:
            continue
        gci = float(r.get("gross_commission") or 0)
        if r.get("entry_type") == "LISTING": listing_actual[m] += gci
        elif r.get("entry_type") == "BUYER": buyer_actual[m] += gci

    # Build output: each month with income, cost_of_sales, operating_expenses, plus computed totals + actual_income overlays
    out_months = []
    yearly_totals = {"income": 0, "cost_of_sales": 0, "operating_expenses": 0, "gross_profit": 0, "net_profit": 0,
                     "actual_listing_income": 0, "actual_buyer_income": 0, "actual_total_income": 0}
    for r in months:
        m = r["month"]
        inc = dict(r.get("income") or {})
        cogs = dict(r.get("cost_of_sales") or {})
        opex = dict(r.get("operating_expenses") or {})
        # Overlay actuals from closed deals (only if user hasn't manually set a value > the actual)
        actual_listing = round(listing_actual.get(m, 0), 2)
        actual_buyer = round(buyer_actual.get(m, 0), 2)
        total_income = sum(float(v or 0) for v in inc.values())
        total_cogs = sum(float(v or 0) for v in cogs.values())
        total_opex = sum(float(v or 0) for v in opex.values())
        gross_profit = total_income - total_cogs
        net_profit = gross_profit - total_opex
        out_months.append({
            "id": r["id"],
            "month": m,
            "income": inc,
            "cost_of_sales": cogs,
            "operating_expenses": opex,
            "notes": r.get("notes"),
            "totals": {
                "income": round(total_income, 2),
                "cost_of_sales": round(total_cogs, 2),
                "operating_expenses": round(total_opex, 2),
                "gross_profit": round(gross_profit, 2),
                "net_profit": round(net_profit, 2),
            },
            "actual_listing_income": actual_listing,
            "actual_buyer_income": actual_buyer,
            "actual_total_income": round(actual_listing + actual_buyer, 2),
        })
        yearly_totals["income"] += total_income
        yearly_totals["cost_of_sales"] += total_cogs
        yearly_totals["operating_expenses"] += total_opex
        yearly_totals["gross_profit"] += gross_profit
        yearly_totals["net_profit"] += net_profit
        yearly_totals["actual_listing_income"] += actual_listing
        yearly_totals["actual_buyer_income"] += actual_buyer
        yearly_totals["actual_total_income"] += actual_listing + actual_buyer
    for k in yearly_totals: yearly_totals[k] = round(yearly_totals[k], 2)
    return {
        "year": year,
        "months": out_months,
        "yearly_totals": yearly_totals,
    }


@router.get("/clients/{client_id}/financials")
def get_financials(client_id: int, request: Request, year: Optional[int] = None):
    workspace_id = _ws(request)
    return _financials_with_actuals(client_id, year or datetime.utcnow().year, workspace_id)


@router.put("/clients/{client_id}/financials/{month}")
def update_financial_month(client_id: int, month: int, payload: FinancialMonthIn, request: Request, year: Optional[int] = None):
    workspace_id = _ws(request)
    yr = year or datetime.utcnow().year
    if not (1 <= month <= 12):
        raise HTTPException(400, "month must be 1-12")
    bundle = _ensure_business_plan(client_id, yr, workspace_id)
    plan_id = bundle["plan"]["id"]
    _ensure_financial_months(plan_id)
    data = payload.dict(exclude_unset=True)
    if not data:
        raise HTTPException(400, "No fields to update")
    res = _supabase.table("monthly_financials").update(data).eq("business_plan_id", plan_id).eq("month", month).execute()
    return res.data[0] if res.data else None


@router.get("/me/financials")
def my_financials(request: Request, year: Optional[int] = None):
    cc = _my_client(request)
    return _financials_with_actuals(cc["id"], year or datetime.utcnow().year, cc["workspace_id"])


@router.put("/me/financials/{month}")
def my_update_financial_month(month: int, payload: FinancialMonthIn, request: Request, year: Optional[int] = None):
    cc = _my_client(request)
    yr = year or datetime.utcnow().year
    if not (1 <= month <= 12):
        raise HTTPException(400, "month must be 1-12")
    bundle = _ensure_business_plan(cc["id"], yr, cc["workspace_id"])
    plan_id = bundle["plan"]["id"]
    _ensure_financial_months(plan_id)
    data = payload.dict(exclude_unset=True)
    if not data:
        raise HTTPException(400, "No fields to update")
    res = _supabase.table("monthly_financials").update(data).eq("business_plan_id", plan_id).eq("month", month).execute()
    return res.data[0] if res.data else None


# ════════════════════════════════════════════════════════════
# Excel Imports (CTE / MREA legacy workbooks)
# ════════════════════════════════════════════════════════════
# Accept an .xlsx upload, detect the workbook type by sheet names, and
# extract: agent metadata + GCI goal + splits + sale prices/comm rates
# (from "Business Plan" tab) and daily lead gen entries (from "Daily
# Lead Gen Entry" tab). Returns a preview; client confirms with
# apply=true to persist.

from fastapi import UploadFile, File, Form  # noqa: E402

def _safe_num(v):
    if v is None: return None
    try:
        f = float(v)
        return f if f == f else None  # NaN check
    except Exception:
        return None


def _parse_cte_workbook(xlsx_bytes: bytes) -> dict:
    """Parse a CTE 2019-style workbook. Returns extracted-data dict."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    sheets = wb.sheetnames

    out = {
        "workbook_type": None,
        "client": {},
        "plan": {},
        "economic": {},
        "budget": {},
        "activity_logs": [],
        "warnings": [],
    }

    is_cte = ("Business Plan" in sheets and "Daily Lead Gen Entry" in sheets)
    if is_cte:
        out["workbook_type"] = "CTE_2019"
    else:
        out["workbook_type"] = "UNKNOWN"
        out["warnings"].append("Sheet names don't match a known template — no Business Plan / Daily Lead Gen Entry tabs found.")
        return out

    # ── Business Plan tab ──
    try:
        bp = wb["Business Plan"]
        # Iterate cells we care about
        # Row 7: D7 = agent name; H7..L7 = Estimated Taxes label + value
        # Row 8: D8 = Type (Ind/Team); L8 = Split to Office
        # Row 9: D9 = Big Why; L9 = Split Cap
        # Row 16+: Lead agent row. C16=name, E16=GCI goal, H16=ASP listing, I16=ASP buyer, J16=comm listing, K16=comm buyer
        rowvals = {}
        for r_idx, row in enumerate(bp.iter_rows(values_only=True), 1):
            if r_idx > 25: break
            rowvals[r_idx] = list(row)
        def cell(r, c):  # 1-indexed columns
            try: return rowvals.get(r, [])[c-1]
            except Exception: return None

        team_or_individual = "TEAM" if (cell(8, 4) and "team" in str(cell(8, 4)).lower()) else "INDIVIDUAL"
        out["client"]["team_or_individual"] = team_or_individual
        team_name = cell(7, 4)
        if team_name and team_or_individual == "TEAM":
            out["client"]["team_name"] = str(team_name)
        big_why = cell(9, 4)
        # If big_why cell is blank, the legacy spreadsheets sometimes had a hint instead.
        if big_why and not str(big_why).startswith("Your Big Why"):
            out["client"]["big_why"] = str(big_why)

        tax_pct = _safe_num(cell(7, 12))   # L7 — fraction (0.17)
        split_to_office = _safe_num(cell(8, 12))  # L8 — fraction (0.30)
        split_cap = _safe_num(cell(9, 12))  # L9 — dollar (21000)
        if tax_pct is not None and tax_pct < 1: out["budget"]["income_tax_pct"] = tax_pct
        if split_cap is not None: out["budget"]["split_cap"] = split_cap
        if split_to_office is not None and split_to_office < 1:
            # We don't store a percent split — instead infer dollar paid_to_brokerage from cap
            # Leave unset so the comp-plan default applies on import
            pass

        # Lead agent row (typically row 16, but find first row with C populated and E populated)
        lead_name = None
        lead_gci = None
        lead_asp_listing = None
        lead_asp_buyer = None
        lead_comm_listing = None
        lead_comm_buyer = None
        for r in range(15, 22):
            name = cell(r, 3); gci = _safe_num(cell(r, 5))
            if name and gci and gci > 1000:
                lead_name = str(name)
                lead_gci = gci
                lead_asp_listing = _safe_num(cell(r, 8))
                lead_asp_buyer = _safe_num(cell(r, 9))
                lead_comm_listing = _safe_num(cell(r, 10))
                lead_comm_buyer = _safe_num(cell(r, 11))
                break
        if lead_gci: out["plan"]["gci_target"] = lead_gci
        if lead_asp_listing: out["economic"]["seller_avg_sale_price"] = lead_asp_listing
        if lead_asp_buyer: out["economic"]["buyer_avg_sale_price"] = lead_asp_buyer
        if lead_comm_listing and lead_comm_listing < 1: out["economic"]["commission_rate_listing"] = lead_comm_listing
        if lead_comm_buyer and lead_comm_buyer < 1: out["economic"]["commission_rate_buyer"] = lead_comm_buyer
        if lead_name:
            out["_extracted_lead_name"] = lead_name
    except Exception as e:
        out["warnings"].append(f"Error reading Business Plan tab: {e}")

    # ── Daily Lead Gen Entry tab ──
    try:
        dle = wb["Daily Lead Gen Entry"]
        # Headers on R3, data from R4+
        # Columns: A=row#, B=Name, C=Date, D=Hours, E=Dials, F=Contacts, G=Nurtures,
        #   H=Listing Appts Set, I=Listing Appts Held, J=*LISTING(s) Signed,
        #   K=Buyer Appts Set, L=Buyer Appts Held
        # Some CTE versions also have M=Buyers Signed
        for r_idx, row in enumerate(dle.iter_rows(values_only=True), 1):
            if r_idx <= 3: continue
            if r_idx > 400: break  # safety
            if not row or len(row) < 12: continue
            dt = row[2]
            if not dt: continue
            # Convert date
            log_date = None
            if hasattr(dt, "isoformat"):
                log_date = dt.date().isoformat() if hasattr(dt, "date") else str(dt)[:10]
            else:
                try:
                    log_date = datetime.strptime(str(dt)[:10], "%Y-%m-%d").date().isoformat()
                except Exception:
                    continue
            entry = {
                "log_date": log_date,
                "hours_prospected": _safe_num(row[3]) or 0,
                "dials": int(_safe_num(row[4]) or 0),
                "contacts_made": int(_safe_num(row[5]) or 0),
                "nurtures": int(_safe_num(row[6]) or 0),
                "listing_appts_set": int(_safe_num(row[7]) or 0),
                "listing_appts_held": int(_safe_num(row[8]) or 0),
                "listings_signed": int(_safe_num(row[9]) or 0),
                "buyer_appts_set": int(_safe_num(row[10]) or 0),
                "buyer_appts_held": int(_safe_num(row[11]) or 0),
            }
            if len(row) > 12 and row[12] is not None:
                entry["buyers_signed"] = int(_safe_num(row[12]) or 0)
            # Skip empty rows (date-only with all zeros)
            non_zero = any(v for k, v in entry.items() if k != "log_date")
            if non_zero:
                out["activity_logs"].append(entry)
    except Exception as e:
        out["warnings"].append(f"Error reading Daily Lead Gen Entry tab: {e}")

    return out


@router.post("/clients/{client_id}/import-cte")
async def import_cte_workbook(
    client_id: int,
    request: Request,
    file: UploadFile = File(...),
    apply: bool = Form(False),
):
    """Upload a CTE / MREA xlsx. With apply=False (default), parses and returns a preview.
    With apply=True, also persists the extracted data into the coaching_client + plan + activity logs."""
    workspace_id = _ws(request)
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Only .xlsx / .xlsm files are supported")
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:  # 10 MB cap
        raise HTTPException(400, "File too large (10 MB max)")
    parsed = _parse_cte_workbook(contents)

    if not apply or parsed.get("workbook_type") == "UNKNOWN":
        return {"preview": True, "parsed": parsed, "summary": {
            "workbook_type": parsed.get("workbook_type"),
            "client_fields_found": len(parsed.get("client") or {}),
            "plan_fields_found": len(parsed.get("plan") or {}),
            "economic_fields_found": len(parsed.get("economic") or {}),
            "budget_fields_found": len(parsed.get("budget") or {}),
            "activity_logs_found": len(parsed.get("activity_logs") or []),
            "warnings": parsed.get("warnings") or [],
        }}

    # ── Apply ──
    yr = datetime.utcnow().year
    bundle = _ensure_business_plan(client_id, yr, workspace_id)
    plan_id = bundle["plan"]["id"]

    if parsed.get("client"):
        _supabase.table("coaching_clients").update(parsed["client"]).eq("id", client_id).execute()
    if parsed.get("plan"):
        _supabase.table("business_plans").update(parsed["plan"]).eq("id", plan_id).execute()
    if parsed.get("economic"):
        _supabase.table("economic_models").update(parsed["economic"]).eq("business_plan_id", plan_id).execute()
    if parsed.get("budget"):
        _supabase.table("budget_models").update(parsed["budget"]).eq("business_plan_id", plan_id).execute()

    # Activity logs — upsert one at a time
    inserted = 0
    skipped = 0
    for log in parsed.get("activity_logs", []):
        try:
            row = {"coaching_client_id": client_id, **log}
            _supabase.table("coaching_activity_logs").upsert(row, on_conflict="coaching_client_id,log_date").execute()
            inserted += 1
        except Exception:
            skipped += 1

    return {
        "applied": True,
        "client_id": client_id,
        "fields_updated": {
            "client": list((parsed.get("client") or {}).keys()),
            "plan": list((parsed.get("plan") or {}).keys()),
            "economic": list((parsed.get("economic") or {}).keys()),
            "budget": list((parsed.get("budget") or {}).keys()),
        },
        "activity_logs_imported": inserted,
        "activity_logs_skipped": skipped,
        "warnings": parsed.get("warnings") or [],
    }


@router.get("/clients/{client_id}/scorecard")
def coach_scorecard(client_id: int, request: Request, year: Optional[int] = None):
    workspace_id = _ws(request)
    yr = year or datetime.utcnow().year
    return _scorecard_full(client_id, yr, workspace_id)


@router.get("/clients/{client_id}/activity-goals")
def coach_activity_goals(client_id: int, request: Request):
    """Coach view of an agent's activity goals."""
    workspace_id = _ws(request)
    yr = datetime.utcnow().year
    bundle = _ensure_business_plan(client_id, yr, workspace_id)
    res = _supabase.table("activity_goals").select("*").eq("business_plan_id", bundle["plan"]["id"]).execute()
    return res.data[0] if res.data else {}


# ─── Coaching Dashboard (coach view across all clients) ───

@router.get("/dashboard")
def coaching_dashboard(request: Request):
    """Aggregate book-of-business: who needs attention this week."""
    workspace_id = _ws(request)
    today = date.today()
    yr = today.year
    days_elapsed = (today - date(yr, 1, 1)).days + 1
    expected_pct = days_elapsed / 365.0

    clients = _db("coaching_clients").select("*").eq("status", "ACTIVE").execute().data or []
    clients = [_enrich_client(c) for c in clients]

    rows = []
    total_gci_goal = 0
    total_gci_ytd = 0
    behind_clients = []
    thin_pipeline = []
    no_recent_activity = []
    upcoming_calls = []
    low_keep = []

    from datetime import timedelta as _td
    cutoff_activity = (today - _td(days=7)).isoformat()
    cutoff_calls_lo = today.isoformat()
    cutoff_calls_hi = (today + _td(days=7)).isoformat()

    for c in clients:
        bundle = _ensure_business_plan(c["id"], yr, workspace_id)
        plan = bundle["plan"]
        gci_target = float(plan.get("gci_target") or 0)
        # YTD
        closed = _supabase.table("pipeline_entries").select("gross_commission,closed_date").eq("coaching_client_id", c["id"]).eq("closed", True).execute().data or []
        gci_ytd = sum(float(r.get("gross_commission") or 0) for r in closed if (r.get("closed_date") or "").startswith(str(yr)))
        total_gci_goal += gci_target
        total_gci_ytd += gci_ytd
        actual_pct = (gci_ytd / gci_target) if gci_target > 0 else 0
        gap_pct = round((actual_pct - expected_pct) * 100, 1)

        # Pipeline 10s+9s count (open)
        open_pipe = _supabase.table("pipeline_entries").select("rating,entry_type").eq("coaching_client_id", c["id"]).eq("closed", False).execute().data or []
        hot_listings = sum(1 for p in open_pipe if p.get("entry_type") == "LISTING" and p.get("rating") in (9, 10))

        # Recent activity
        last_act = _supabase.table("coaching_activity_logs").select("log_date").eq("coaching_client_id", c["id"]).order("log_date", desc=True).limit(1).execute().data or []
        last_log_date = last_act[0]["log_date"] if last_act else None
        days_since_log = None
        if last_log_date:
            try:
                d = datetime.strptime(last_log_date, "%Y-%m-%d").date()
                days_since_log = (today - d).days
            except Exception:
                pass

        # Upcoming + last call
        next_call = _supabase.table("coaching_calls").select("id,scheduled_at,call_type,status").eq("coaching_client_id", c["id"]).gte("scheduled_at", cutoff_calls_lo).lte("scheduled_at", cutoff_calls_hi).order("scheduled_at").limit(1).execute().data or []
        last_completed = _supabase.table("coaching_calls").select("commitment_keep_score").eq("coaching_client_id", c["id"]).eq("status", "COMPLETED").order("completed_at", desc=True).limit(1).execute().data or []
        last_keep = last_completed[0]["commitment_keep_score"] if last_completed and last_completed[0].get("commitment_keep_score") is not None else None

        lead = c.get("lead") or {}
        client_summary = {
            "id": c["id"],
            "name": lead.get("name") or "—",
            "comp_plan": c.get("lpt_comp_plan"),
            "cadence": c.get("call_cadence"),
            "gci_target": gci_target,
            "gci_ytd": round(gci_ytd, 2),
            "pace_gap_pct": gap_pct,
            "hot_listings": hot_listings,
            "open_pipeline_total": len(open_pipe),
            "days_since_log": days_since_log,
            "next_call": next_call[0] if next_call else None,
            "last_commit_keep": last_keep,
        }
        rows.append(client_summary)

        if gap_pct < -10:
            behind_clients.append(client_summary)
        if hot_listings == 0 and len(open_pipe) > 0:
            thin_pipeline.append(client_summary)
        elif len(open_pipe) == 0:
            thin_pipeline.append(client_summary)
        if days_since_log is None or days_since_log > 3:
            no_recent_activity.append(client_summary)
        if next_call:
            upcoming_calls.append(client_summary)
        if last_keep is not None and last_keep < 70:
            low_keep.append(client_summary)

    return {
        "totals": {
            "active_clients": len(clients),
            "aggregate_gci_goal": round(total_gci_goal, 2),
            "aggregate_gci_ytd": round(total_gci_ytd, 2),
            "aggregate_pace_pct": round((total_gci_ytd / total_gci_goal * 100), 1) if total_gci_goal else 0,
            "expected_pct": round(expected_pct * 100, 1),
        },
        "all_clients": sorted(rows, key=lambda r: r["pace_gap_pct"]),
        "behind_pace": sorted(behind_clients, key=lambda r: r["pace_gap_pct"]),
        "thin_pipeline": thin_pipeline,
        "no_recent_activity": no_recent_activity,
        "upcoming_calls": upcoming_calls,
        "low_commitment_keep": low_keep,
    }


# ─── GPS (1-3-5) ───

class GPSGoalIn(BaseModel):
    goal_text: Optional[str] = None
    target_number: Optional[float] = None


class GPSPriorityIn(BaseModel):
    priority_text: Optional[str] = None
    target_number: Optional[float] = None
    sort_order: Optional[int] = None


class GPSStrategyIn(BaseModel):
    strategy_text: Optional[str] = None
    target_number: Optional[float] = None
    source_or_method: Optional[str] = None
    sort_order: Optional[int] = None


@router.get("/clients/{client_id}/gps")
def get_gps(client_id: int, request: Request, year: Optional[int] = None):
    workspace_id = _ws(request)
    yr = year or datetime.utcnow().year
    bundle = _ensure_business_plan(client_id, yr, workspace_id)
    plan_id = bundle["plan"]["id"]
    goal_resp = _supabase.table("gps_goals").select("*").eq("business_plan_id", plan_id).execute()
    if goal_resp.data:
        goal = goal_resp.data[0]
    else:
        # Auto-create with the GCI target as the goal
        gci = bundle["plan"].get("gci_target") or 0
        ins = _supabase.table("gps_goals").insert({
            "business_plan_id": plan_id,
            "goal_text": f"Earn ${int(float(gci)):,} in GCI in {yr}",
            "target_number": gci,
        }).execute()
        goal = ins.data[0]
    priorities = _supabase.table("gps_priorities").select("*").eq("gps_goal_id", goal["id"]).order("sort_order").execute().data or []
    out_priorities = []
    for p in priorities:
        strats = _supabase.table("gps_strategies").select("*").eq("gps_priority_id", p["id"]).order("sort_order").execute().data or []
        p["strategies"] = strats
        out_priorities.append(p)
    return {"goal": goal, "priorities": out_priorities}


@router.patch("/gps-goals/{goal_id}")
def update_gps_goal(goal_id: int, payload: GPSGoalIn, request: Request):
    data = payload.dict(exclude_unset=True)
    return _supabase.table("gps_goals").update(data).eq("id", goal_id).execute().data[0]


@router.post("/gps-goals/{goal_id}/priorities")
def create_priority(goal_id: int, payload: GPSPriorityIn, request: Request):
    row = {"gps_goal_id": goal_id, **payload.dict(exclude_unset=True)}
    return _supabase.table("gps_priorities").insert(row).execute().data[0]


@router.patch("/gps-priorities/{pid}")
def update_priority(pid: int, payload: GPSPriorityIn, request: Request):
    return _supabase.table("gps_priorities").update(payload.dict(exclude_unset=True)).eq("id", pid).execute().data[0]


@router.delete("/gps-priorities/{pid}")
def delete_priority(pid: int, request: Request):
    _supabase.table("gps_priorities").delete().eq("id", pid).execute()
    return {"ok": True}


@router.post("/gps-priorities/{pid}/strategies")
def create_strategy(pid: int, payload: GPSStrategyIn, request: Request):
    row = {"gps_priority_id": pid, **payload.dict(exclude_unset=True)}
    return _supabase.table("gps_strategies").insert(row).execute().data[0]


@router.patch("/gps-strategies/{sid}")
def update_strategy(sid: int, payload: GPSStrategyIn, request: Request):
    return _supabase.table("gps_strategies").update(payload.dict(exclude_unset=True)).eq("id", sid).execute().data[0]


@router.delete("/gps-strategies/{sid}")
def delete_strategy(sid: int, request: Request):
    _supabase.table("gps_strategies").delete().eq("id", sid).execute()
    return {"ok": True}


# ─── 4-1-1 ───

class FourOneOneItemIn(BaseModel):
    items: list  # [{text, target_number, completed}]


@router.get("/clients/{client_id}/four-one-one")
def get_411(client_id: int, request: Request, period_type: str = "ANNUAL", period_key: Optional[str] = None, year: Optional[int] = None):
    workspace_id = _ws(request)
    yr = year or datetime.utcnow().year
    bundle = _ensure_business_plan(client_id, yr, workspace_id)
    plan_id = bundle["plan"]["id"]

    if not period_key:
        if period_type == "ANNUAL":
            period_key = str(yr)
        elif period_type == "MONTHLY":
            period_key = f"{yr}-{datetime.utcnow().month:02d}"
        else:  # WEEKLY
            iso = datetime.utcnow().isocalendar()
            period_key = f"{iso[0]}-W{iso[1]:02d}"

    rows = _supabase.table("four_one_ones").select("*").eq("business_plan_id", plan_id).eq("period_type", period_type).eq("period_key", period_key).execute().data or []
    by_col = {r["column_key"]: r for r in rows}
    out = {}
    for col in ("JOB", "BUSINESS", "PERSONAL_FINANCIAL", "PERSONAL"):
        out[col] = by_col.get(col, {"items": [], "column_key": col, "period_type": period_type, "period_key": period_key, "business_plan_id": plan_id})
    # Suggestions for ANNUAL Business column from Big Rocks
    if period_type == "ANNUAL" and not out["BUSINESS"].get("items"):
        em = bundle["economic_model"]
        gci = float(bundle["plan"].get("gci_target") or 0)
        econ = calc_economic(gci, em)
        out["BUSINESS"]["suggestions"] = [
            {"text": "Listings taken", "target_number": econ["listings_taken"]["value"]},
            {"text": "Buyers shown", "target_number": econ["buyers_shown"]["value"]},
            {"text": "Listing appts", "target_number": econ["listing_appts_annual"]["value"]},
            {"text": "Buyer consults", "target_number": econ["buyer_consults_annual"]["value"]},
            {"text": "GCI", "target_number": gci},
        ]
    return {"period_type": period_type, "period_key": period_key, "columns": out}


@router.put("/clients/{client_id}/four-one-one")
def upsert_411(client_id: int, request: Request, period_type: str, period_key: str, column_key: str, payload: FourOneOneItemIn, year: Optional[int] = None):
    workspace_id = _ws(request)
    yr = year or datetime.utcnow().year
    bundle = _ensure_business_plan(client_id, yr, workspace_id)
    plan_id = bundle["plan"]["id"]
    row = {
        "business_plan_id": plan_id,
        "period_type": period_type,
        "period_key": period_key,
        "column_key": column_key,
        "items": payload.items,
    }
    return _supabase.table("four_one_ones").upsert(row, on_conflict="business_plan_id,period_type,period_key,column_key").execute().data[0]


# ─── HybridShare / Recruiting ───

LPT_HYBRIDSHARE_TIERS = [
    {"tier": 1, "pct_pool": 31, "min_active": 1,  "max_per_bp": 2325, "max_per_bb": 775},
    {"tier": 2, "pct_pool": 18, "min_active": 4,  "max_per_bp": 1350, "max_per_bb": 450},
    {"tier": 3, "pct_pool": 7,  "min_active": 8,  "max_per_bp": 525,  "max_per_bb": 175},
    {"tier": 4, "pct_pool": 7,  "min_active": 12, "max_per_bp": 525,  "max_per_bb": 175},
    {"tier": 5, "pct_pool": 7,  "min_active": 16, "max_per_bp": 525,  "max_per_bb": 175},
    {"tier": 6, "pct_pool": 10, "min_active": 19, "max_per_bp": 750,  "max_per_bb": 250},
    {"tier": 7, "pct_pool": 20, "min_active": 20, "max_per_bp": 1500, "max_per_bb": 500},
]
LPT_BP_MAX_TOTAL = 7500
LPT_BB_MAX_TOTAL = 2500

LPT_PERFORMANCE_AWARDS = [
    {"award": "White Badge",  "txns": 1,  "shares_bp": 50,   "shares_bb": 25},
    {"award": "Silver Badge", "txns": 3,  "shares_bp": 50,   "shares_bb": 25},
    {"award": "Gold Badge",   "txns": 15, "shares_bp": 600,  "shares_bb": 300},
    {"award": "Black Badge",  "txns": 35, "shares_bp": 1800, "shares_bb": None},  # BP only
]


class RecruitIn(BaseModel):
    recruit_name: str
    recruit_email: Optional[str] = None
    recruit_phone: Optional[str] = None
    status: Optional[str] = "HITLIST"
    tier: Optional[int] = None
    sponsor_recruit_id: Optional[int] = None
    co_sponsor_recruit_id: Optional[int] = None
    join_date: Optional[str] = None
    comp_plan: Optional[str] = None
    current_ytd_core_txns: Optional[int] = None
    last_contact_date: Optional[str] = None
    source: Optional[str] = None
    notes: Optional[str] = None


class RecruitUpdate(BaseModel):
    recruit_name: Optional[str] = None
    recruit_email: Optional[str] = None
    recruit_phone: Optional[str] = None
    status: Optional[str] = None
    tier: Optional[int] = None
    sponsor_recruit_id: Optional[int] = None
    co_sponsor_recruit_id: Optional[int] = None
    join_date: Optional[str] = None
    comp_plan: Optional[str] = None
    current_ytd_core_txns: Optional[int] = None
    last_contact_date: Optional[str] = None
    source: Optional[str] = None
    notes: Optional[str] = None


@router.get("/clients/{client_id}/recruits")
def list_recruits(client_id: int, request: Request, status: Optional[str] = None):
    qb = _db("coaching_recruits").select("*").eq("coaching_client_id", client_id)
    if status:
        qb = qb.eq("status", status)
    return qb.order("status").order("created_at", desc=True).execute().data or []


@router.post("/clients/{client_id}/recruits")
def create_recruit(client_id: int, payload: RecruitIn, request: Request):
    workspace_id = _ws(request)
    row = {"workspace_id": workspace_id, "coaching_client_id": client_id, "tier": payload.tier or 1, **payload.dict(exclude_unset=True)}
    return _supabase.table("coaching_recruits").insert(row).execute().data[0]


@router.patch("/recruits/{rid}")
def update_recruit(rid: int, payload: RecruitUpdate, request: Request):
    data = payload.dict(exclude_unset=True)
    if not data:
        raise HTTPException(400, "No fields")
    return _supabase.table("coaching_recruits").update(data).eq("id", rid).execute().data[0]


@router.delete("/recruits/{rid}")
def delete_recruit(rid: int, request: Request):
    _supabase.table("coaching_recruits").delete().eq("id", rid).execute()
    return {"ok": True}


@router.get("/clients/{client_id}/hybridshare")
def hybridshare_summary(client_id: int, request: Request):
    """Computed HybridShare summary for a coaching client. Returns:
    - Tier table with current count, lock state, tier subtotal
    - Performance awards progress for the agent + each signed recruit
    - 5-year freedom-number projection (params: assume_recruits_per_year, pct_bp, etc.)
    """
    cc = _supabase.table("coaching_clients").select("*").eq("id", client_id).single().execute().data
    if not cc:
        raise HTTPException(404, "Client not found")
    own_plan = cc.get("lpt_comp_plan")  # BUSINESS_BUILDER or BROKERAGE_PARTNER

    recruits = _supabase.table("coaching_recruits").select("*").eq("coaching_client_id", client_id).execute().data or []
    signed = [r for r in recruits if r.get("status") == "SIGNED"]

    # Group by tier
    by_tier = {}
    for r in signed:
        t = r.get("tier") or 1
        by_tier.setdefault(t, []).append(r)

    tiers_out = []
    total_projected_income = 0
    for tdef in LPT_HYBRIDSHARE_TIERS:
        t = tdef["tier"]
        slots = by_tier.get(t, [])
        bp_count = sum(1 for r in slots if r.get("comp_plan") == "BROKERAGE_PARTNER")
        bb_count = sum(1 for r in slots if r.get("comp_plan") == "BUSINESS_BUILDER")
        total_active_in_tier = bp_count + bb_count
        unlocked = total_active_in_tier >= tdef["min_active"]
        tier_subtotal = bp_count * tdef["max_per_bp"] + bb_count * tdef["max_per_bb"]
        if unlocked:
            total_projected_income += tier_subtotal
        tiers_out.append({
            **tdef,
            "bp_count": bp_count,
            "bb_count": bb_count,
            "total_active": total_active_in_tier,
            "unlocked": unlocked,
            "tier_subtotal": tier_subtotal,
            "slots": slots,
        })

    # Performance Awards for the agent
    agent_txns = 0
    # Sum closed pipeline for the agent in current year
    yr = datetime.utcnow().year
    closed = _supabase.table("pipeline_entries").select("closed_date").eq("coaching_client_id", client_id).eq("closed", True).execute().data or []
    agent_txns = sum(1 for r in closed if (r.get("closed_date") or "").startswith(str(yr)))

    awards_progress = []
    for a in LPT_PERFORMANCE_AWARDS:
        shares = a["shares_bp"] if own_plan == "BROKERAGE_PARTNER" else a["shares_bb"]
        if shares is None:
            achieved = False
            progress_pct = 0
        else:
            achieved = agent_txns >= a["txns"]
            progress_pct = min(100, round((agent_txns / a["txns"]) * 100, 1))
        awards_progress.append({
            "award": a["award"],
            "txns_required": a["txns"],
            "txns_current": agent_txns,
            "shares": shares,
            "achieved": achieved,
            "progress_pct": progress_pct,
            "available_to_agent": shares is not None,
        })

    return {
        "comp_plan": own_plan,
        "tiers": tiers_out,
        "total_projected_income_at_full_cap": total_projected_income,
        "max_possible_total": LPT_BP_MAX_TOTAL if own_plan == "BROKERAGE_PARTNER" else LPT_BB_MAX_TOTAL,
        "agent_ytd_txns": agent_txns,
        "performance_awards": awards_progress,
    }


@router.get("/clients/{client_id}/hybridshare/projection")
def hybridshare_projection(
    client_id: int, request: Request,
    recruits_per_year: float = 4.0,
    pct_bp: float = 0.5,
    cap_hit_rate: float = 0.30,
    children_per_recruit: float = 1.5,
):
    """5-year stacked projection. Each year the network grows by recruits_per_year tier-1,
    plus children_per_recruit per existing tier-1 recruit (which fills tier 2, etc.)."""
    cc = _supabase.table("coaching_clients").select("lpt_comp_plan").eq("id", client_id).single().execute().data or {}
    own_plan = cc.get("lpt_comp_plan") or "BROKERAGE_PARTNER"
    pct_bb = 1.0 - pct_bp

    network = [0] * 8  # tier 1..7 counts, ignore index 0
    network[1] = 0
    rows = []
    for yr in range(1, 6):
        # New tier-1 from agent's direct recruiting
        network[1] += recruits_per_year
        # Trickle: each existing tier-N produces children at tier N+1
        for t in range(7, 1, -1):
            parents_at_prev = network[t - 1]
            new_children = parents_at_prev * children_per_recruit if t == 2 else parents_at_prev * (children_per_recruit / 2)  # diminishing
            network[t] += new_children

        # Project income at cap_hit_rate
        income = 0
        unlocked_tiers = 0
        for tdef in LPT_HYBRIDSHARE_TIERS:
            t = tdef["tier"]
            n = network[t]
            if n >= tdef["min_active"]:
                unlocked_tiers += 1
            bp_count = n * pct_bp
            bb_count = n * pct_bb
            tier_max = bp_count * tdef["max_per_bp"] + bb_count * tdef["max_per_bb"]
            income += tier_max * cap_hit_rate

        rows.append({
            "year": yr,
            "network_size": int(sum(network[1:8])),
            "tiers_unlocked": unlocked_tiers,
            "projected_income": round(income, 2),
            "tier_breakdown": {f"tier_{t}": int(network[t]) for t in range(1, 8)},
        })

    return {
        "params": {"recruits_per_year": recruits_per_year, "pct_bp": pct_bp, "cap_hit_rate": cap_hit_rate, "children_per_recruit": children_per_recruit},
        "comp_plan": own_plan,
        "years": rows,
    }


# Need to import is_valid_email from main lazily
def is_valid_email(value):
    from main import is_valid_email as _ive
    return _ive(value)
